from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q, F
from django.views.decorators.http import require_POST
from django.db import IntegrityError
from functools import wraps
import datetime
import json

from .models import (
    Service, Worker, Patient, Visit, Procedure, ProcedureDate, UserAccount,
    VisitService, QueueNumber, ScheduledVisit, get_next_queue_number,
    ClinicalProcedure, Expense, VisitAttachment, ReferralPartner,
    Glasses, GlassSale, EyeRefraction, EyeKRT, EyeIOP,
    EyeVisualAcuity, EyeExamForm, DischargeForm,
    LasikForm, StrabismusForm, PostOpForm,
    LabVisit, LabVisitService, LabQueueNumber, get_next_lab_queue_number,
    LabServiceTemplate,
    OptikaPatientPrescription, OptikaSale,
)


# ─── PHONE NORMALIZATION ───

def normalize_phone(phone):
    """Strip +998 prefix and spaces, keep just the digits after country code."""
    phone = phone.strip().replace(' ', '').replace('-', '')
    if phone.startswith('+998'):
        phone = phone[4:]
    elif phone.startswith('998'):
        phone = phone[3:]
    return phone


# ─── AUTH HELPERS ───

def get_current_user(request):
    """Get UserAccount from session, or a dict for system accounts."""
    user_id = request.session.get('user_id')
    if not user_id:
        return None
    # System accounts (admin, receptionist, laboratory)
    if isinstance(user_id, str) and user_id.startswith('system_'):
        role = request.session.get('user_role', '')
        name = request.session.get('user_name', '')
        # Return a simple object with role attribute
        class SystemUser:
            def __init__(self, r, n):
                self.role = r
                self.name = n
                self.full_name = n
                self.pk = user_id
        return SystemUser(role, name)
    try:
        return UserAccount.objects.get(pk=user_id)
    except (UserAccount.DoesNotExist, ValueError, TypeError):
        return None


def login_required_custom(view_func):
    """Decorator: redirect to login if no session."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.session.get('user_id'):
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper


def role_required(*roles):
    """Decorator: check that logged-in user has one of the specified roles."""
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            session_role = request.session.get('user_role')
            if not session_role:
                return redirect('login')
            if session_role not in roles:
                return redirect('login')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# ─── LOGIN / LOGOUT ───

HARDCODED_ACCOUNTS = {
    'admin': {'username': 'sherzod', 'password': '911666222', 'role': 'admin', 'display_name': 'Sherzod'},
    'receptionist': {'username': 'reception', 'password': '123', 'role': 'receptionist', 'display_name': 'Reception'},
    'laboratory': {'username': 'labaratoriya', 'password': '123', 'role': 'laboratory', 'display_name': 'Laboratoriya'},
    'seller': {'username': 'optika', 'password': '123', 'role': 'seller', 'display_name': 'Optika'},
}


def login_view(request):
    error = ''
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        password = request.POST.get('phone', '').strip()

        # Check hardcoded system accounts first
        name_lower = name.lower()
        for key, acc in HARDCODED_ACCOUNTS.items():
            if name_lower == acc['username'] and password == acc['password']:
                request.session['user_id'] = f'system_{key}'
                request.session['user_role'] = acc['role']
                request.session['user_name'] = acc['display_name']
                if acc['role'] == 'admin':
                    return redirect('admin_panel')
                elif acc['role'] == 'receptionist':
                    return redirect('receptionist_panel')
                elif acc['role'] == 'laboratory':
                    return redirect('lab_panel')
                elif acc['role'] == 'seller':
                    return redirect('seller_sales')
                break
        else:
            # Check DB-based accounts (doctors, sellers)
            phone = normalize_phone(password)
            try:
                user = UserAccount.objects.get(phone=phone)
                full_input = name_lower
                full_name = f"{user.name} {user.surname}".lower()
                if full_input == full_name or full_input == user.name.lower():
                    request.session['user_id'] = user.pk
                    request.session['user_role'] = user.role
                    request.session['user_name'] = user.full_name
                    if user.role == 'doctor':
                        request.session['doctor_id'] = user.worker_id
                        return redirect('doctor_panel', doctor_id=user.worker_id)
                    elif user.role == 'seller':
                        return redirect('seller_sales')
                    elif user.role == 'admin':
                        return redirect('admin_panel')
                    elif user.role == 'receptionist':
                        return redirect('receptionist_panel')
                else:
                    error = 'Ism yoki parol noto\'g\'ri'
            except UserAccount.DoesNotExist:
                error = 'Ism yoki parol noto\'g\'ri'
            if not error:
                pass  # Already redirected
            else:
                pass  # error is already set
    return render(request, 'login.html', {'error': error})


def logout_view(request):
    request.session.flush()
    return redirect('login')


# ─── MAIN ADMIN PANEL ───

@role_required('admin')
def admin_panel(request):
    user = get_current_user(request)
    services = Service.objects.all()
    workers = Worker.objects.filter(worker_type='doctor').prefetch_related('services')
    expenses = Expense.objects.all()
    referral_partners = ReferralPartner.objects.all()
    glasses = Glasses.objects.all()
    return render(request, 'admin/panel.html', {
        'services': services,
        'workers': workers,
        'all_services': services,
        'expenses': expenses,
        'referral_partners': referral_partners,
        'glasses': glasses,
        'user': user,
    })


@role_required('admin')
def add_service(request):
    if request.method == 'POST':
        Service.objects.create(
            name=request.POST['name'],
            description=request.POST.get('description', ''),
            price=int(request.POST.get('price', 0) or 0),
        )
    return redirect('admin_panel')


@role_required('admin')
def edit_service(request, pk):
    service = get_object_or_404(Service, pk=pk)
    if request.method == 'POST':
        service.name = request.POST['name']
        service.description = request.POST.get('description', '')
        service.price = int(request.POST.get('price', 0) or 0)
        service.save()
    return redirect('admin_panel')


@role_required('admin')
def delete_service(request, pk):
    get_object_or_404(Service, pk=pk).delete()
    return redirect('admin_panel')


@role_required('admin')
def add_worker(request):
    if request.method == 'POST':
        phone = normalize_phone(request.POST['phone'])
        if Worker.objects.filter(phone=phone).exists():
            services = Service.objects.all()
            workers = Worker.objects.filter(worker_type='doctor').prefetch_related('services')
            expenses = Expense.objects.all()
            return render(request, 'admin/panel.html', {
                'services': services, 'workers': workers, 'all_services': services,
                'expenses': expenses,
                'error': 'Bu telefon raqami allaqachon mavjud',
            })
        if UserAccount.objects.filter(phone=phone).exists():
            services = Service.objects.all()
            workers = Worker.objects.filter(worker_type='doctor').prefetch_related('services')
            expenses = Expense.objects.all()
            return render(request, 'admin/panel.html', {
                'services': services, 'workers': workers, 'all_services': services,
                'expenses': expenses,
                'error': 'Bu telefon raqami boshqa foydalanuvchida mavjud',
            })
        worker = Worker.objects.create(
            name=request.POST['name'],
            surname=request.POST['surname'],
            phone=phone,
            salary=int(request.POST.get('salary', 0) or 0),
            worker_type='doctor',
        )
        service_ids = request.POST.getlist('services')
        if service_ids:
            worker.services.set(service_ids)
            worker.service_id = service_ids[0]
            worker.save()
        # Auto-create UserAccount for doctors
        UserAccount.objects.filter(phone=worker.phone).exclude(worker=worker).exclude(role='admin').delete()
        UserAccount.objects.update_or_create(
            worker=worker,
            defaults={
                'name': worker.name,
                'surname': worker.surname,
                'phone': worker.phone,
                'role': 'doctor',
            }
        )
    return redirect('admin_panel')


@role_required('admin')
def edit_worker(request, pk):
    worker = get_object_or_404(Worker, pk=pk)
    if request.method == 'POST':
        phone = normalize_phone(request.POST['phone'])
        if Worker.objects.filter(phone=phone).exclude(pk=pk).exists():
            services = Service.objects.all()
            workers = Worker.objects.filter(worker_type='doctor').prefetch_related('services')
            expenses = Expense.objects.all()
            return render(request, 'admin/panel.html', {
                'services': services, 'workers': workers, 'all_services': services,
                'expenses': expenses,
                'error': 'Bu telefon raqami allaqachon mavjud',
            })
        dup_account = UserAccount.objects.filter(phone=phone).exclude(worker=worker).exists()
        if dup_account:
            services = Service.objects.all()
            workers = Worker.objects.filter(worker_type='doctor').prefetch_related('services')
            expenses = Expense.objects.all()
            return render(request, 'admin/panel.html', {
                'services': services, 'workers': workers, 'all_services': services,
                'expenses': expenses,
                'error': 'Bu telefon raqami boshqa foydalanuvchida mavjud',
            })
        worker.name = request.POST['name']
        worker.surname = request.POST['surname']
        worker.phone = phone
        worker.salary = int(request.POST.get('salary', 0) or 0)
        worker.worker_type = 'doctor'
        service_ids = request.POST.getlist('services')
        worker.service_id = service_ids[0] if service_ids else None
        worker.save()
        worker.services.set(service_ids)
        # Sync UserAccount
        UserAccount.objects.filter(phone=worker.phone).exclude(worker=worker).exclude(role='admin').delete()
        UserAccount.objects.update_or_create(
            worker=worker,
            defaults={
                'name': worker.name,
                'surname': worker.surname,
                'phone': worker.phone,
                'role': 'doctor',
            }
        )
    return redirect('admin_panel')


@role_required('admin')
def delete_worker(request, pk):
    worker = get_object_or_404(Worker, pk=pk)
    UserAccount.objects.filter(worker=worker).delete()
    worker.delete()
    return redirect('admin_panel')


# ─── CLINICAL PROCEDURE CATALOG ───

@role_required('admin')
def add_clinical_procedure(request):
    if request.method == 'POST':
        ClinicalProcedure.objects.create(
            name=request.POST['name'],
            description=request.POST.get('description', ''),
            price_per_session=int(request.POST.get('price_per_session', 0) or 0),
        )
    return redirect('admin_panel')


@role_required('admin')
def edit_clinical_procedure(request, pk):
    cp = get_object_or_404(ClinicalProcedure, pk=pk)
    if request.method == 'POST':
        cp.name = request.POST['name']
        cp.description = request.POST.get('description', '')
        cp.price_per_session = int(request.POST.get('price_per_session', 0) or 0)
        cp.save()
    return redirect('admin_panel')


@role_required('admin')
def delete_clinical_procedure(request, pk):
    get_object_or_404(ClinicalProcedure, pk=pk).delete()
    return redirect('admin_panel')


# ─── EXPENSES ───

@role_required('admin')
def add_expense(request):
    if request.method == 'POST':
        Expense.objects.create(
            name=request.POST['name'],
            amount=int(request.POST.get('amount', 0) or 0),
            expense_type=request.POST.get('expense_type', 'recurring'),
            date=request.POST.get('date') or None,
        )
    return redirect('admin_panel')


@role_required('admin')
def edit_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        expense.name = request.POST['name']
        expense.amount = int(request.POST.get('amount', 0) or 0)
        expense.expense_type = request.POST.get('expense_type', expense.expense_type)
        expense.date = request.POST.get('date') or None
        expense.save()
    return redirect('admin_panel')


@role_required('admin')
def delete_expense(request, pk):
    get_object_or_404(Expense, pk=pk).delete()
    return redirect('admin_panel')


# ─── REFERRAL PARTNERS ───

@role_required('admin')
def add_referral_partner(request):
    if request.method == 'POST':
        commission_type = request.POST.get('commission_type', 'percent')
        ReferralPartner.objects.create(
            name=request.POST['name'],
            surname=request.POST['surname'],
            phone=normalize_phone(request.POST.get('phone', '')),
            clinic_name=request.POST.get('clinic_name', ''),
            commission_type=commission_type,
            commission_percent=int(request.POST.get('commission_percent', '10') or '10'),
            commission_fixed=int(request.POST.get('commission_fixed', '0') or '0'),
            notes=request.POST.get('notes', ''),
        )
    return redirect('admin_panel')


@role_required('admin')
def edit_referral_partner(request, pk):
    rp = get_object_or_404(ReferralPartner, pk=pk)
    if request.method == 'POST':
        rp.name = request.POST['name']
        rp.surname = request.POST['surname']
        rp.phone = normalize_phone(request.POST.get('phone', ''))
        rp.clinic_name = request.POST.get('clinic_name', '')
        rp.commission_type = request.POST.get('commission_type', 'percent')
        rp.commission_percent = int(request.POST.get('commission_percent', '10') or '10')
        rp.commission_fixed = int(request.POST.get('commission_fixed', '0') or '0')
        rp.notes = request.POST.get('notes', '')
        rp.is_active = 'is_active' in request.POST
        rp.save()
    return redirect('admin_panel')


@role_required('admin')
def delete_referral_partner(request, pk):
    get_object_or_404(ReferralPartner, pk=pk).delete()
    return redirect('admin_panel')


@role_required('admin')
def analytics_referrals(request):
    period = request.GET.get('period', 'month')
    date_from = _get_date_filter(period)

    partners = ReferralPartner.objects.all()
    partner_data = []

    for rp in partners:
        visits = Visit.objects.filter(
            referred_by=rp, status__in=['done', 'has_procedure']
        )
        if date_from:
            visits = visits.filter(created_at__gte=date_from)
        count = visits.count()
        revenue = sum(v.final_price for v in visits)
        commission = sum(v.commission_amount for v in visits)
        partner_data.append({
            'id': rp.pk,
            'name': rp.full_name,
            'clinic': rp.clinic_name or 'Mustaqil',
            'percent': rp.commission_display,
            'patients': count,
            'revenue': revenue,
            'commission': commission,
            'is_active': rp.is_active,
        })

    partner_data.sort(key=lambda x: x['commission'], reverse=True)
    total_commission = sum(p['commission'] for p in partner_data)
    total_revenue = sum(p['revenue'] for p in partner_data)

    return render(request, 'admin/analytics_referrals.html', {
        'period': period,
        'partner_data': partner_data,
        'total_commission': total_commission,
        'total_revenue': total_revenue,
        'partner_data_json': json.dumps(partner_data),
    })


# ─── ANALYTICS ───

def _get_date_filter(period):
    now = timezone.now()
    if period == 'week':
        return now - datetime.timedelta(days=7)
    elif period == 'month':
        return now - datetime.timedelta(days=30)
    elif period == '6m':
        return now - datetime.timedelta(days=180)
    elif period in ('year', '1y'):
        return now - datetime.timedelta(days=365)
    return None


def _get_period_months(period):
    """Number of months in the selected period (for salary-based spendings)."""
    if period == 'month':
        return 1
    elif period == '6m':
        return 6
    elif period in ('year', '1y'):
        return 12
    else:
        first = Visit.objects.order_by('created_at').first()
        if first:
            return max(1, (timezone.now() - first.created_at).days // 30)
        return 1


def _get_expense_spendings(period, date_from=None):
    """Calculate spendings from Expense model."""
    period_months = _get_period_months(period)
    recurring_total = Expense.objects.filter(
        expense_type='recurring'
    ).aggregate(total=Sum('amount'))['total'] or 0
    recurring_spendings = recurring_total * period_months
    one_time_qs = Expense.objects.filter(expense_type='one_time')
    if date_from:
        one_time_qs = one_time_qs.filter(date__gte=date_from)
    one_time_spendings = one_time_qs.aggregate(total=Sum('amount'))['total'] or 0
    return recurring_spendings + one_time_spendings


def _get_monthly_trends(period='6m'):
    """Compute trends for all metrics (revenue, patients, spendings, net) based on period."""
    now = timezone.now()
    trends = []
    recurring_monthly = Expense.objects.filter(
        expense_type='recurring'
    ).aggregate(total=Sum('amount'))['total'] or 0

    if period == 'month':
        daily_expense = round(recurring_monthly / 30)
        for i in range(29, -1, -1):
            day = now - datetime.timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + datetime.timedelta(days=1)
            visits = Visit.objects.filter(
                status__in=['done', 'has_procedure'],
                created_at__gte=day_start, created_at__lt=day_end,
            )
            rev = sum(v.final_price for v in visits)
            pts = visits.values('patient').distinct().count()
            one_time_day = Expense.objects.filter(
                expense_type='one_time', date=day_start.date()
            ).aggregate(total=Sum('amount'))['total'] or 0
            day_spendings = daily_expense + one_time_day
            trends.append({
                'label': day_start.strftime('%d %b'),
                'revenue': rev,
                'patients': pts,
                'spendings': day_spendings,
                'net_revenue': rev - day_spendings,
            })
    else:
        if period == '1y':
            num_months = 12
        elif period == 'all':
            num_months = 24
        else:
            num_months = 6

        for i in range(num_months - 1, -1, -1):
            dt = now - datetime.timedelta(days=i * 30)
            month_start = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1)
            visits = Visit.objects.filter(
                status__in=['done', 'has_procedure'],
                created_at__gte=month_start, created_at__lt=month_end,
            )
            rev = sum(v.final_price for v in visits)
            pts = visits.values('patient').distinct().count()
            one_time_month = Expense.objects.filter(
                expense_type='one_time',
                date__gte=month_start.date(), date__lt=month_end.date()
            ).aggregate(total=Sum('amount'))['total'] or 0
            month_spendings = recurring_monthly + one_time_month
            trends.append({
                'label': month_start.strftime('%b %Y'),
                'revenue': rev,
                'patients': pts,
                'spendings': month_spendings,
                'net_revenue': rev - month_spendings,
            })
    return trends


@role_required('admin')
def analytics_overall(request):
    period = request.GET.get('period', 'month')
    selected_metric = request.GET.get('metric', 'net_revenue')
    date_from = _get_date_filter(period)

    visits = Visit.objects.all()
    if date_from:
        visits = visits.filter(created_at__gte=date_from)

    completed_visits = visits.filter(status__in=['done', 'has_procedure'])

    # Revenue from medical visits
    medical_revenue = sum(v.final_price for v in completed_visits)

    # Lab revenue
    lab_visits_qs = LabVisit.objects.filter(status='done')
    if date_from:
        lab_visits_qs = lab_visits_qs.filter(created_at__gte=date_from)
    lab_revenue = sum(lv.total_price for lv in lab_visits_qs)

    # Total revenue (medical + lab only — optika is separate)
    revenue = medical_revenue + lab_revenue

    # Referral commissions
    total_commission = sum(v.commission_amount for v in completed_visits)

    # Spendings (from Expense model)
    spendings = _get_expense_spendings(period, date_from)
    net_revenue = revenue - spendings - total_commission

    # Patients
    total_patients = completed_visits.values('patient').distinct().count()

    # Ratings
    rated_visits = visits.filter(rated=True)
    avg_doctor = rated_visits.aggregate(avg=Avg('doctor_rating'))['avg']
    avg_clinic = rated_visits.aggregate(avg=Avg('clinic_rating'))['avg']
    avg_doctor_rating = round(avg_doctor) if avg_doctor else None
    avg_clinic_rating = round(avg_clinic) if avg_clinic else None

    returning_patients_count = 0
    patient_ids = completed_visits.values_list('patient_id', flat=True).distinct()
    for pid in patient_ids:
        if Visit.objects.filter(patient_id=pid).count() > 1:
            returning_patients_count += 1
    returning_pct = round((returning_patients_count / total_patients * 100)) if total_patients > 0 else 0

    patient_list = []
    for v in visits.select_related('patient', 'doctor', 'service').order_by('-created_at'):
        patient_list.append({
            'patient_id': v.patient_id,
            'name': v.patient.full_name,
            'phone': v.patient.phone,
            'doctor': v.doctor.full_name if v.doctor else "O'chirilgan doktor",
            'service': v.service.name if v.service else "—",
            'date': v.created_at,
            'discount': v.discount_amount,
            'discount_reason': v.discount_reason,
            'has_discount': v.discount_amount > 0,
        })

    # Trends with all 4 metrics (uses same period as top filter)
    trends = _get_monthly_trends(period)

    return render(request, 'admin/analytics_overall.html', {
        'period': period,
        'revenue': revenue,
        'spendings': spendings,
        'net_revenue': net_revenue,
        'total_commission': total_commission,
        'total_patients': total_patients,
        'avg_doctor_rating': avg_doctor_rating,
        'avg_clinic_rating': avg_clinic_rating,
        'returning_patients': returning_patients_count,
        'returning_pct': returning_pct,
        'patient_list': patient_list,
        'trends_json': json.dumps(trends),
        'selected_metric': selected_metric,
    })


@role_required('admin')
def analytics_services(request):
    period = request.GET.get('period', 'month')
    date_from = _get_date_filter(period)

    services = Service.objects.all()
    service_data = []

    for s in services:
        visits = Visit.objects.filter(
            visit_services__service=s, status__in=['done', 'has_procedure']
        ).distinct()
        if date_from:
            visits = visits.filter(created_at__gte=date_from)
        count = visits.count()
        # Revenue from this service specifically (sum price_at_time from VisitService rows)
        rev = VisitService.objects.filter(
            service=s,
            visit__status__in=['done', 'has_procedure'],
            **(dict(visit__created_at__gte=date_from) if date_from else {})
        ).aggregate(total=Sum('price_at_time'))['total'] or 0
        service_data.append({
            'name': s.name,
            'count': count,
            'revenue': rev,
        })

    service_data.sort(key=lambda x: x['revenue'], reverse=True)

    return render(request, 'admin/analytics_services.html', {
        'period': period,
        'service_data': service_data,
        'service_data_json': json.dumps(service_data),
    })


@role_required('admin')
def analytics_doctors(request):
    period = request.GET.get('period', 'month')
    date_from = _get_date_filter(period)

    doctors = Worker.objects.filter(worker_type='doctor')
    doctor_data = []

    for d in doctors:
        visits = Visit.objects.filter(doctor=d, status__in=['done', 'has_procedure'])
        if date_from:
            visits = visits.filter(created_at__gte=date_from)
        count = visits.count()
        rev = sum(v.final_price for v in visits)
        rated = visits.filter(rated=True)
        avg_r = rated.aggregate(avg=Avg('doctor_rating'))['avg']
        avg_rating = round(avg_r) if avg_r else None
        doctor_data.append({
            'name': d.full_name,
            'patients_seen': count,
            'revenue': rev,
            'rating': avg_rating,
        })

    doctor_data.sort(key=lambda x: x['revenue'], reverse=True)

    return render(request, 'admin/analytics_doctors.html', {
        'period': period,
        'doctor_data': doctor_data,
        'doctor_data_json': json.dumps(doctor_data),
    })


@role_required('admin')
def admin_patient_detail(request, patient_id):
    """Admin view: full detail of a patient with all visits and procedures."""
    patient = get_object_or_404(Patient, pk=patient_id)
    visits = Visit.objects.filter(patient=patient).select_related(
        'doctor', 'service'
    ).prefetch_related('procedures__dates').order_by('-created_at')

    total_spent = 0
    for v in visits.filter(status__in=['done', 'has_procedure']):
        total_spent += v.final_price

    return render(request, 'admin/patient_detail.html', {
        'patient': patient,
        'visits': visits,
        'total_spent': total_spent,
    })


# ─── RECEPTIONIST PANEL ───

def _get_receptionist_badge_counts():
    """Get notification and procedure counts for receptionist sidebar badges."""
    # Notifications: done + unrated (regardless of procedure completion)
    notification_count = Visit.objects.filter(status='done', rated=False).count()

    # Procedures: upcoming dates (today or tomorrow, not completed, not yet notified)
    today = timezone.localdate()
    tomorrow = today + datetime.timedelta(days=1)
    # Reset overdue notified (7+ days past scheduled, still not completed)
    ProcedureDate.objects.filter(
        completed=False, notified=True,
        scheduled_date__lt=today - datetime.timedelta(days=7),
    ).update(notified=False)
    procedure_count = ProcedureDate.objects.filter(
        completed=False, notified=False, scheduled_date__lte=tomorrow
    ).count()

    return notification_count, procedure_count


@login_required_custom
def receptionist_panel(request):
    user = get_current_user(request)
    if user.role not in ('receptionist', 'admin'):
        return redirect('login')

    search = request.GET.get('search', '')
    patients = Patient.objects.all().order_by('-created_at')
    if search:
        search_terms = search.strip().split()
        if len(search_terms) >= 2:
            patients = patients.filter(
                Q(phone__icontains=search) |
                Q(name__icontains=search) |
                Q(surname__icontains=search) |
                (Q(name__icontains=search_terms[0]) & Q(surname__icontains=search_terms[1])) |
                (Q(name__icontains=search_terms[1]) & Q(surname__icontains=search_terms[0]))
            )
        else:
            patients = patients.filter(
                Q(phone__icontains=search) |
                Q(name__icontains=search) |
                Q(surname__icontains=search)
            )

    doctors = Worker.objects.filter(worker_type='doctor').prefetch_related('services')
    services = Service.objects.all()

    # Get waiting visits (to show edit button)
    waiting_visits = Visit.objects.filter(
        status='waiting',
    ).select_related('patient', 'doctor').prefetch_related('visit_services__service')
    waiting_map = {}
    for wv in waiting_visits:
        waiting_map[wv.patient_id] = wv

    # Get waiting lab visits
    waiting_lab_visits = LabVisit.objects.filter(
        status='waiting',
    ).select_related('patient').prefetch_related('lab_services', 'queue')

    # Scheduled visits (not yet activated) — includes both doctor and lab
    today = timezone.localdate()
    scheduled_visits = ScheduledVisit.objects.filter(
        activated=False, scheduled_date__gte=today,
    ).select_related('patient', 'doctor').order_by('scheduled_date', 'scheduled_time')[:20]

    notif_count, proc_count = _get_receptionist_badge_counts()

    today_str = timezone.now().strftime('%Y-%m-%d')
    now_time_str = timezone.now().strftime('%H:%M')

    # Scheduled visits for today: patient_id -> sv.pk (for quick activation button)
    today_sv_by_patient = {}
    for sv in ScheduledVisit.objects.filter(scheduled_date=today, activated=False).select_related('doctor'):
        doctor_name = 'Laboratoriya' if sv.is_lab else (sv.doctor.full_name if sv.doctor else '')
        today_sv_by_patient[sv.patient_id] = {'pk': sv.pk, 'doctor_name': doctor_name}

    referral_partners = ReferralPartner.objects.filter(is_active=True)

    waiting_total = len(waiting_map) + waiting_lab_visits.count()

    return render(request, 'receptionist/panel.html', {
        'patients': patients,
        'doctors': doctors,
        'services': services,
        'search': search,
        'user': user,
        'notif_count': notif_count,
        'proc_count': proc_count,
        'waiting_map': waiting_map,
        'waiting_lab_visits': waiting_lab_visits,
        'waiting_total': waiting_total,
        'scheduled_visits': scheduled_visits,
        'today_str': today_str,
        'now_time_str': now_time_str,
        'today_sv_by_patient': today_sv_by_patient,
        'referral_partners': referral_partners,
    })


@login_required_custom
def add_patient(request):
    if request.method == 'POST':
        phone = normalize_phone(request.POST['phone'])
        patient, created = Patient.objects.get_or_create(
            phone=phone,
            defaults={
                'name': request.POST['name'],
                'surname': request.POST['surname'],
            }
        )
        if not created:
            patient.name = request.POST['name']
            patient.surname = request.POST['surname']

        # Receptionist fills birthdate and gender
        gender = request.POST.get('gender', '')
        if gender:
            patient.gender = gender
        bd = request.POST.get('birth_date', '')
        if bd:
            try:
                patient.birth_date = datetime.datetime.strptime(bd, '%Y-%m-%d').date()
            except ValueError:
                pass
        patient.save()

        # Check if this is a laboratory visit
        destination = request.POST.get('destination', 'doctor')
        if destination == 'laboratory':
            lab_service_ids = request.POST.getlist('lab_services')

            # Check if scheduling for future
            visit_date_str = request.POST.get('visit_date', '')
            today = timezone.localdate()
            visit_date = today
            if visit_date_str:
                try:
                    visit_date = datetime.datetime.strptime(visit_date_str, '%Y-%m-%d').date()
                except ValueError:
                    visit_date = today

            if visit_date > today:
                # Future lab appointment: store as ScheduledVisit with is_lab=True
                lab_data = json.dumps(lab_service_ids)
                ScheduledVisit.objects.create(
                    patient=patient,
                    doctor=None,
                    scheduled_date=visit_date,
                    is_lab=True,
                    lab_services_data=lab_data,
                )
                return redirect(reverse('receptionist_panel') + '?success=1')

            # Create a lab visit for today
            lab_visit = LabVisit.objects.create(patient=patient, status='waiting')
            # Add selected lab services from DB template
            svc_map = {t.number: t for t in LabServiceTemplate.objects.filter(number__in=[int(x) for x in lab_service_ids])}
            for ls_id in lab_service_ids:
                tpl = svc_map.get(int(ls_id))
                if tpl:
                    LabVisitService.objects.create(
                        lab_visit=lab_visit,
                        service_number=tpl.number,
                        service_name=tpl.name,
                        price=tpl.price,
                    )
            # Lab queue number (resets daily)
            try:
                lqn = get_next_lab_queue_number()
                LabQueueNumber.objects.create(lab_visit=lab_visit, number=lqn, date=timezone.localdate())
            except Exception:
                pass
            return redirect('receptionist_print_lab_ticket', lab_visit_id=lab_visit.pk)

        doctor = Worker.objects.get(pk=request.POST['doctor'])
        discount = int(request.POST.get('discount_amount', '0') or '0')
        discount_reason = request.POST.get('discount_reason', '')

        # Check if scheduling for future
        visit_date_str = request.POST.get('visit_date', '')
        today = timezone.localdate()
        visit_date = today
        if visit_date_str:
            try:
                visit_date = datetime.datetime.strptime(visit_date_str, '%Y-%m-%d').date()
            except ValueError:
                visit_date = today

        if visit_date > today:
            # Future date: create ScheduledVisit
            ScheduledVisit.objects.create(
                patient=patient,
                doctor=doctor,
                scheduled_date=visit_date,
                notes=discount_reason,
            )
            return redirect(reverse('receptionist_panel') + '?success=1')

        # Referral partner
        referral_id = request.POST.get('referred_by', '')
        referred_by = None
        if referral_id:
            try:
                referred_by = ReferralPartner.objects.get(pk=int(referral_id))
            except (ReferralPartner.DoesNotExist, ValueError):
                pass

        # Today: create Visit immediately
        first_svc = doctor.services.first()
        visit = Visit.objects.create(
            patient=patient,
            doctor=doctor,
            doctor_name=doctor.full_name,
            referred_by=referred_by,
            service=first_svc,
            discount_amount=discount,
            discount_reason=discount_reason,
            status='waiting',
        )
        # Create VisitService records (multi-service)
        service_ids = request.POST.getlist('services')
        if service_ids:
            for sid in service_ids:
                svc = Service.objects.get(pk=sid)
                VisitService.objects.create(visit=visit, service=svc, price_at_time=svc.price)
        elif first_svc:
            VisitService.objects.create(visit=visit, service=first_svc, price_at_time=first_svc.price)
        # Queue number (per-doctor)
        qnum = get_next_queue_number(doctor=doctor)
        QueueNumber.objects.create(visit=visit, doctor=doctor, number=qnum, date=timezone.localdate())
        return redirect('receptionist_print_ticket', visit_id=visit.pk)
    return redirect('receptionist_panel')


@login_required_custom
def edit_patient(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        patient.name = request.POST.get('name', patient.name)
        patient.surname = request.POST.get('surname', patient.surname)
        raw_phone = request.POST.get('phone', '')
        if raw_phone:
            patient.phone = normalize_phone(raw_phone)
        patient.save()
    return redirect('receptionist_panel')


@login_required_custom
def delete_patient(request, pk):
    get_object_or_404(Patient, pk=pk).delete()
    return redirect('receptionist_panel')


@login_required_custom
def redirect_patient(request, pk):
    """
    Redirect a patient to a doctor.
    If the patient has a 'has_procedure' visit for this doctor, reuse it
    (set back to 'waiting') instead of creating a new visit.
    """
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        destination = request.POST.get('destination', 'doctor')

        # Handle lab destination
        if destination == 'laboratory':
            lab_service_ids = request.POST.getlist('lab_services')

            # Check if scheduling for future
            visit_date_str = request.POST.get('visit_date', '')
            today = timezone.localdate()
            visit_date = today
            if visit_date_str:
                try:
                    visit_date = datetime.datetime.strptime(visit_date_str, '%Y-%m-%d').date()
                except ValueError:
                    visit_date = today

            if visit_date > today:
                # Future lab appointment
                lab_data = json.dumps(lab_service_ids)
                ScheduledVisit.objects.create(
                    patient=patient,
                    doctor=None,
                    scheduled_date=visit_date,
                    is_lab=True,
                    lab_services_data=lab_data,
                )
                return redirect(reverse('receptionist_panel') + '?success=1')

            lab_visit = LabVisit.objects.create(patient=patient, status='waiting')
            if lab_service_ids:
                svc_map2 = {t.number: t for t in LabServiceTemplate.objects.filter(number__in=[int(x) for x in lab_service_ids])}
                for ls_id in lab_service_ids:
                    tpl = svc_map2.get(int(ls_id))
                    if tpl:
                        LabVisitService.objects.create(
                            lab_visit=lab_visit,
                            service_number=tpl.number,
                            service_name=tpl.name,
                            price=tpl.price,
                        )
            # Lab queue number (resets daily)
            try:
                lqn = get_next_lab_queue_number()
                LabQueueNumber.objects.create(lab_visit=lab_visit, number=lqn, date=timezone.localdate())
            except Exception:
                pass
            return redirect('receptionist_print_lab_ticket', lab_visit_id=lab_visit.pk)

        # Handle doctor destination
        doctor_id = request.POST.get('doctor', '')
        if not doctor_id:
            return redirect('receptionist_panel')
        doctor = Worker.objects.get(pk=doctor_id)
        discount = int(request.POST.get('discount_amount', '0') or '0')
        discount_reason = request.POST.get('discount_reason', '')

        # Check if scheduling for future
        visit_date_str = request.POST.get('visit_date', '')
        today = timezone.localdate()
        visit_date = today
        if visit_date_str:
            try:
                visit_date = datetime.datetime.strptime(visit_date_str, '%Y-%m-%d').date()
            except ValueError:
                visit_date = today

        if visit_date > today:
            ScheduledVisit.objects.create(
                patient=patient, doctor=doctor,
                scheduled_date=visit_date,
                notes=discount_reason,
            )
            return redirect(reverse('receptionist_panel') + '?success=1')

        # Referral partner
        referral_id = request.POST.get('referred_by', '')
        referred_by = None
        if referral_id:
            try:
                referred_by = ReferralPartner.objects.get(pk=int(referral_id))
            except (ReferralPartner.DoesNotExist, ValueError):
                pass

        # Check if a specific procedure was selected
        existing_procedure_id = request.POST.get('existing_procedure', '')

        # Check if there is an existing has_procedure visit for this patient+doctor
        existing_proc_visit = None
        if existing_procedure_id:
            try:
                proc = Procedure.objects.get(pk=int(existing_procedure_id))
                if proc.visit.patient == patient and proc.visit.status == 'has_procedure':
                    existing_proc_visit = proc.visit
            except (Procedure.DoesNotExist, ValueError):
                pass

        if not existing_proc_visit:
            existing_proc_visit = Visit.objects.filter(
                patient=patient, doctor=doctor, status='has_procedure',
            ).first()

        first_svc = doctor.services.first()
        service_ids = request.POST.getlist('services')
        if existing_proc_visit:
            existing_proc_visit.status = 'waiting'
            if referred_by:
                existing_proc_visit.referred_by = referred_by
            existing_proc_visit.save()
            visit = existing_proc_visit
        else:
            visit = Visit.objects.create(
                patient=patient, doctor=doctor,
                doctor_name=doctor.full_name,
                referred_by=referred_by,
                service=first_svc,
                discount_amount=discount,
                discount_reason=discount_reason,
                status='waiting',
            )
            if service_ids:
                for sid in service_ids:
                    svc = Service.objects.get(pk=sid)
                    VisitService.objects.create(visit=visit, service=svc, price_at_time=svc.price)
            elif not existing_procedure_id and first_svc:
                VisitService.objects.create(visit=visit, service=first_svc, price_at_time=first_svc.price)
        # Queue number (per-doctor, for both new and reactivated)
        if not hasattr(visit, 'queue') or visit.queue is None:
            try:
                qnum = get_next_queue_number(doctor=doctor)
                QueueNumber.objects.create(visit=visit, doctor=doctor, number=qnum, date=timezone.localdate())
            except Exception:
                pass
        return redirect('receptionist_print_ticket', visit_id=visit.pk)
    return redirect('receptionist_panel')


@login_required_custom
def receptionist_notifications(request):
    """
    Show visits where:
    - status is 'done'
    - visit has NOT been rated yet
    Even if there are incomplete procedures, the visit appears here for rating.
    """
    notification_visits = Visit.objects.filter(
        status='done', rated=False
    ).select_related('patient', 'doctor', 'service').order_by('-completed_at')

    notif_count, proc_count = _get_receptionist_badge_counts()

    return render(request, 'receptionist/notifications.html', {
        'visits': notification_visits,
        'notif_count': notif_count,
        'proc_count': proc_count,
    })


@login_required_custom
def rate_visit(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    if request.method == 'POST':
        doctor_rating = int(request.POST.get('doctor_rating', 0))
        clinic_rating = int(request.POST.get('clinic_rating', 0))
        visit.doctor_rating = max(1, min(10, doctor_rating))
        visit.clinic_rating = max(1, min(10, clinic_rating))
        visit.rated = True
        visit.save()
    return redirect('receptionist_notifications')


@login_required_custom
def receptionist_procedures(request):
    today = timezone.localdate()
    tomorrow = today + datetime.timedelta(days=1)

    procedure_dates = ProcedureDate.objects.filter(
        completed=False,
        scheduled_date__lte=tomorrow,
    ).select_related(
        'procedure__patient', 'procedure__doctor', 'procedure__visit'
    ).order_by('scheduled_date')

    overdue_dates = ProcedureDate.objects.filter(
        completed=False,
        notified=True,
        scheduled_date__lt=today - datetime.timedelta(days=7),
    ).select_related('procedure__patient', 'procedure__doctor')

    for od in overdue_dates:
        od.notified = False
        od.save()

    notif_count, proc_count = _get_receptionist_badge_counts()

    return render(request, 'receptionist/procedures.html', {
        'procedure_dates': procedure_dates,
        'today': today,
        'notif_count': notif_count,
        'proc_count': proc_count,
    })


@login_required_custom
def notify_procedure(request, pk):
    pd = get_object_or_404(ProcedureDate, pk=pk)
    pd.notified = True
    pd.save()
    return redirect('receptionist_procedures')


# ─── DOCTOR PANEL ───

@login_required_custom
def doctor_panel(request, doctor_id):
    user = get_current_user(request)
    if user.role not in ('doctor', 'admin'):
        return redirect('login')
    # Doctor can only access their own panel (admin can access any)
    if user.role == 'doctor' and user.worker_id != doctor_id:
        return redirect('login')

    doctor = get_object_or_404(Worker, pk=doctor_id, worker_type='doctor')
    period = request.GET.get('period', 'today')
    tab = request.GET.get('tab', 'current')

    # Current patients: group by patient, not by individual visit
    # A "current" patient is one who has any waiting/in_progress visit with this doctor
    current_visits = Visit.objects.filter(
        doctor=doctor,
        status__in=['waiting', 'in_progress']
    ).select_related('patient', 'service').order_by('-created_at')

    # Group current visits by patient
    current_patients_map = {}
    for v in current_visits:
        pid = v.patient_id
        if pid not in current_patients_map:
            current_patients_map[pid] = {
                'patient': v.patient,
                'latest_visit': v,
                'visit_count': Visit.objects.filter(patient=v.patient, doctor=doctor).count(),
            }

    current_patients = list(current_patients_map.values())

    # Done visits filter
    done_filter = Visit.objects.filter(
        doctor=doctor,
        status__in=['done', 'has_procedure']
    ).select_related('patient', 'service')

    now = timezone.now()
    local_today = timezone.localdate()
    if period == 'today':
        done_filter = done_filter.filter(completed_at__date=local_today)
    elif period == 'week':
        done_filter = done_filter.filter(completed_at__gte=now - datetime.timedelta(days=7))
    elif period == 'month':
        done_filter = done_filter.filter(completed_at__gte=now - datetime.timedelta(days=30))

    done_visits = done_filter.order_by('-completed_at')

    current_count = len(current_patients)
    done_count = done_visits.count()

    # Build JSON data for edit modal (all fields for each done visit)
    done_visits_edit_data = {}
    for v in done_visits:
        done_visits_edit_data[v.pk] = {
            'tashxis': v.tashxis or '',
            'tavsiya': v.tavsiya or '',
            'anamnesis_morbi': v.anamnesis_morbi or '',
            'operatsiyalar': v.operatsiyalar or '',
            'tomizilgan': v.tomizilgan or '',
            'koz_oynak_kontakt_linza': v.koz_oynak_kontakt_linza or '',
            'anamnesis_vitae': v.anamnesis_vitae or '',
            'allergiya': v.allergiya or '',
            'qovoqlar': v.qovoqlar_koz_yosh_yollari or '',
            'koz_soqqasi': v.koz_soqqasi or '',
            'koz_olmasi': v.koz_olmasi or '',
            'gavhar': v.gavhar or '',
            'korish_otkirligi': v.korish_otkirligi or '',
            'shishasimon_tana': v.shishasimon_tana or '',
            'fundus_od': v.fundus_od or '',
            'fundus_os': v.fundus_os or '',
            'ubm_od': v.ubm_od or '',
            'ubm_os': v.ubm_os or '',
            'oct_od': v.oct_od or '',
            'oct_os': v.oct_os or '',
            'kp_od': v.kp_od or '',
            'kp_os': v.kp_os or '',
            'b_skan_od': v.b_skan_od or '',
            'b_skan_os': v.b_skan_os or '',
            'periferiya_od': v.periferiya_od or '',
            'periferiya_os': v.periferiya_os or '',
            'retsept_uzoq_od': v.retsept_uzoq_od or '',
            'retsept_uzoq_os': v.retsept_uzoq_os or '',
            'retsept_yaqin_od': v.retsept_yaqin_od or '',
            'retsept_yaqin_os': v.retsept_yaqin_os or '',
            'retsept_kontakt_od': v.retsept_kontakt_od or '',
            'retsept_kontakt_os': v.retsept_kontakt_os or '',
            'oftalmoskopiya': v.oftalmoskopiya or '',
            'sklera': v.sklera or '',
            'shox_parda': v.shox_parda or '',
            'old_kamera': v.old_kamera or '',
            'rangdor_parda': v.rangdor_parda_qorachiq or '',
        }

    return render(request, 'doctor/panel.html', {
        'doctor': doctor,
        'current_patients': current_patients,
        'current_visits': current_visits,
        'done_visits': done_visits,
        'current_count': current_count,
        'done_count': done_count,
        'period': period,
        'tab': tab,
        'user': user,
        'done_visits_edit_json': json.dumps(done_visits_edit_data),
    })


@login_required_custom
def doctor_patient_detail(request, doctor_id, patient_id):
    """
    Doctor views a patient: shows ALL visits and ALL procedures for this patient,
    not just one visit.
    """
    user = get_current_user(request)
    if user.role not in ('doctor', 'admin'):
        return redirect('login')

    doctor = get_object_or_404(Worker, pk=doctor_id, worker_type='doctor')
    patient = get_object_or_404(Patient, pk=patient_id)

    # All visits for this patient with this doctor
    visits = Visit.objects.filter(
        patient=patient, doctor=doctor
    ).select_related('service').prefetch_related(
        'procedures__dates', 'visit_attachments'
    ).order_by('-created_at')

    # The current active visit (waiting/in_progress) if any
    active_visit = visits.filter(status__in=['waiting', 'in_progress']).first()

    # If there is a waiting visit, set it to in_progress
    if active_visit and active_visit.status == 'waiting':
        active_visit.status = 'in_progress'
        active_visit.save()

    # All procedures across all visits for this patient (with this doctor)
    all_procedures = Procedure.objects.filter(
        patient=patient, doctor=doctor
    ).prefetch_related('dates').order_by('-created_at')

    # ALL visits for this patient across ALL doctors (for unified history)
    all_visits = Visit.objects.filter(
        patient=patient
    ).select_related('doctor', 'service').prefetch_related(
        'procedures__dates', 'visit_attachments', 'refractions', 'krts', 'iops',
        'visual_acuities', 'visit_services__service',
    ).order_by('-created_at')

    # Patient-level forms (for the hujjatlar section, grouped by visit)
    eye_exam_forms = patient.eye_exam_forms.select_related('doctor').order_by('-created_at')
    discharge_forms = patient.discharge_forms.select_related('doctor').order_by('-created_at')
    lasik_forms = patient.lasik_forms.select_related('doctor').order_by('-created_at')
    strabismus_forms = patient.strabismus_forms.select_related('doctor').order_by('-created_at')
    postop_forms = patient.postop_forms.select_related('doctor').order_by('-created_at')

    # Group forms by visit_id for per-visit display
    forms_by_visit = {}
    for e in eye_exam_forms:
        forms_by_visit.setdefault(e.visit_id, []).append({'type': 'eye_exam', 'label': "Ko'z ko'rigi", 'obj': e})
    for d in discharge_forms:
        forms_by_visit.setdefault(d.visit_id, []).append({'type': 'discharge', 'label': "Ko'chirma", 'obj': d})
    for l in lasik_forms:
        forms_by_visit.setdefault(l.visit_id, []).append({'type': 'lasik', 'label': 'LASIK', 'obj': l})
    for s in strabismus_forms:
        forms_by_visit.setdefault(s.visit_id, []).append({'type': 'strabismus', 'label': "Ko'squlilik", 'obj': s})
    for p in postop_forms:
        forms_by_visit.setdefault(p.visit_id, []).append({'type': 'postop', 'label': "Oper. ko'chirma", 'obj': p})

    # Referral partners for doctor to assign
    referral_partners = ReferralPartner.objects.filter(is_active=True)

    # Auto-fill data for LASIK form from latest IOP and REF
    latest_iop = EyeIOP.objects.filter(
        visit__patient=patient
    ).order_by('-created_at').first()
    latest_ref = EyeRefraction.objects.filter(
        visit__patient=patient
    ).order_by('-created_at').first()

    # Tibbiy yozuv tashxis from active visit (used as source for 3 new forms)
    active_tashxis = active_visit.tashxis if active_visit else ''

    lasik_autofill = {
        'vgd_od': latest_iop.od_value if latest_iop else '',
        'vgd_os': latest_iop.os_value if latest_iop else '',
        'preop_sph_od': latest_ref.od_s if latest_ref else '',
        'preop_cyl_od': latest_ref.od_c if latest_ref else '',
        'preop_axis_od': latest_ref.od_a if latest_ref else '',
        'preop_sph_os': latest_ref.os_s if latest_ref else '',
        'preop_cyl_os': latest_ref.os_c if latest_ref else '',
        'preop_axis_os': latest_ref.os_a if latest_ref else '',
    }

    # Build JSON data for edit modal
    visits_edit_data = {}
    for v in all_visits:
        visits_edit_data[v.pk] = {
            'tashxis': v.tashxis or '',
            'tavsiya': v.tavsiya or '',
            'anamnesis_morbi': v.anamnesis_morbi or '',
            'operatsiyalar': v.operatsiyalar or '',
            'tomizilgan': v.tomizilgan or '',
            'koz_oynak_kontakt_linza': v.koz_oynak_kontakt_linza or '',
            'anamnesis_vitae': v.anamnesis_vitae or '',
            'allergiya': v.allergiya or '',
            'qovoqlar': v.qovoqlar_koz_yosh_yollari or '',
            'koz_soqqasi': v.koz_soqqasi or '',
            'koz_olmasi': v.koz_olmasi or '',
            'gavhar': v.gavhar or '',
            'korish_otkirligi': v.korish_otkirligi or '',
            'shishasimon_tana': v.shishasimon_tana or '',
            'fundus_od': v.fundus_od or '',
            'fundus_os': v.fundus_os or '',
            'ubm_od': v.ubm_od or '',
            'ubm_os': v.ubm_os or '',
            'oct_od': v.oct_od or '',
            'oct_os': v.oct_os or '',
            'kp_od': v.kp_od or '',
            'kp_os': v.kp_os or '',
            'b_skan_od': v.b_skan_od or '',
            'b_skan_os': v.b_skan_os or '',
            'periferiya_od': v.periferiya_od or '',
            'periferiya_os': v.periferiya_os or '',
            'retsept_uzoq_od': v.retsept_uzoq_od or '',
            'retsept_uzoq_os': v.retsept_uzoq_os or '',
            'retsept_yaqin_od': v.retsept_yaqin_od or '',
            'retsept_yaqin_os': v.retsept_yaqin_os or '',
            'retsept_kontakt_od': v.retsept_kontakt_od or '',
            'retsept_kontakt_os': v.retsept_kontakt_os or '',
            'oftalmoskopiya': v.oftalmoskopiya or '',
            'sklera': v.sklera or '',
            'shox_parda': v.shox_parda or '',
            'old_kamera': v.old_kamera or '',
            'rangdor_parda': v.rangdor_parda_qorachiq or '',
        }

    return render(request, 'doctor/patient_detail.html', {
        'doctor': doctor,
        'patient': patient,
        'visits': visits,
        'active_visit': active_visit,
        'all_procedures': all_procedures,
        'all_visits': all_visits,
        'referral_partners': referral_partners,
        'eye_exam_forms': eye_exam_forms,
        'discharge_forms': discharge_forms,
        'lasik_forms': lasik_forms,
        'strabismus_forms': strabismus_forms,
        'postop_forms': postop_forms,
        'lasik_autofill': lasik_autofill,
        'active_tashxis': active_tashxis,
        'visits_edit_json': json.dumps(visits_edit_data),
        'forms_by_visit': forms_by_visit,
    })


@login_required_custom
def doctor_patient_full_history(request, doctor_id, patient_id):
    """Full patient history across ALL doctors."""
    user = get_current_user(request)
    if user.role not in ('doctor', 'admin'):
        return redirect('login')

    doctor = get_object_or_404(Worker, pk=doctor_id, worker_type='doctor')
    patient = get_object_or_404(Patient, pk=patient_id)

    visits = list(Visit.objects.filter(
        patient=patient
    ).select_related(
        'doctor', 'service', 'referred_by',
    ).prefetch_related(
        'visit_services__service', 'procedures__dates', 'visit_attachments',
        'refractions', 'krts', 'iops', 'visual_acuities',
    ).order_by('-created_at'))

    # Build JSON data for edit modal
    visits_edit_data = {}
    for v in visits:
        visits_edit_data[v.pk] = {
            'tashxis': v.tashxis or '',
            'tavsiya': v.tavsiya or '',
            'anamnesis_morbi': v.anamnesis_morbi or '',
            'operatsiyalar': v.operatsiyalar or '',
            'tomizilgan': v.tomizilgan or '',
            'koz_oynak_kontakt_linza': v.koz_oynak_kontakt_linza or '',
            'anamnesis_vitae': v.anamnesis_vitae or '',
            'allergiya': v.allergiya or '',
            'qovoqlar': v.qovoqlar_koz_yosh_yollari or '',
            'koz_soqqasi': v.koz_soqqasi or '',
            'koz_olmasi': v.koz_olmasi or '',
            'gavhar': v.gavhar or '',
            'korish_otkirligi': v.korish_otkirligi or '',
            'shishasimon_tana': v.shishasimon_tana or '',
            'fundus_od': v.fundus_od or '',
            'fundus_os': v.fundus_os or '',
            'ubm_od': v.ubm_od or '',
            'ubm_os': v.ubm_os or '',
            'oct_od': v.oct_od or '',
            'oct_os': v.oct_os or '',
            'kp_od': v.kp_od or '',
            'kp_os': v.kp_os or '',
            'b_skan_od': v.b_skan_od or '',
            'b_skan_os': v.b_skan_os or '',
            'periferiya_od': v.periferiya_od or '',
            'periferiya_os': v.periferiya_os or '',
            'retsept_uzoq_od': v.retsept_uzoq_od or '',
            'retsept_uzoq_os': v.retsept_uzoq_os or '',
            'retsept_yaqin_od': v.retsept_yaqin_od or '',
            'retsept_yaqin_os': v.retsept_yaqin_os or '',
            'retsept_kontakt_od': v.retsept_kontakt_od or '',
            'retsept_kontakt_os': v.retsept_kontakt_os or '',
            'oftalmoskopiya': v.oftalmoskopiya or '',
            'sklera': v.sklera or '',
            'shox_parda': v.shox_parda or '',
            'old_kamera': v.old_kamera or '',
            'rangdor_parda': v.rangdor_parda_qorachiq or '',
        }

    eye_exam_forms = EyeExamForm.objects.filter(patient=patient).select_related('doctor').order_by('created_at')
    discharge_forms = DischargeForm.objects.filter(patient=patient).select_related('doctor').order_by('created_at')
    lasik_forms = LasikForm.objects.filter(patient=patient).select_related('doctor').order_by('created_at')
    strabismus_forms = StrabismusForm.objects.filter(patient=patient).select_related('doctor').order_by('created_at')
    postop_forms = PostOpForm.objects.filter(patient=patient).select_related('doctor').order_by('created_at')

    # Group all forms by visit_id and attach to visit objects for template access
    forms_by_visit = {}
    for e in eye_exam_forms:
        forms_by_visit.setdefault(e.visit_id, []).append({'type': 'eye_exam', 'label': "Ko'z ko'rigi", 'obj': e})
    for d in discharge_forms:
        forms_by_visit.setdefault(d.visit_id, []).append({'type': 'discharge', 'label': "Ko'chirma", 'obj': d})
    for l in lasik_forms:
        forms_by_visit.setdefault(l.visit_id, []).append({'type': 'lasik', 'label': 'LASIK', 'obj': l})
    for s in strabismus_forms:
        forms_by_visit.setdefault(s.visit_id, []).append({'type': 'strabismus', 'label': "Ko'squlilik", 'obj': s})
    for p in postop_forms:
        forms_by_visit.setdefault(p.visit_id, []).append({'type': 'postop', 'label': "Oper. ko'chirma", 'obj': p})

    # Attach forms directly to visit objects so template can iterate them
    for v in visits:
        v.visit_forms_list = forms_by_visit.get(v.pk, [])

    return render(request, 'doctor/patient_full_history.html', {
        'doctor': doctor,
        'patient': patient,
        'visits': visits,
        'visits_edit_json': json.dumps(visits_edit_data),
    })


@login_required_custom
def doctor_visit_detail(request, doctor_id, visit_id):
    """Legacy visit detail view - redirects to patient detail."""
    doctor = get_object_or_404(Worker, pk=doctor_id)
    visit = get_object_or_404(Visit, pk=visit_id, doctor=doctor)
    procedures = Procedure.objects.filter(visit=visit).prefetch_related('dates')

    # All visits for this patient with this doctor (for history)
    all_visits = Visit.objects.filter(
        patient=visit.patient, doctor=doctor
    ).select_related('service').prefetch_related('procedures__dates').order_by('-created_at')

    return render(request, 'doctor/visit_detail.html', {
        'doctor': doctor,
        'visit': visit,
        'patient': visit.patient,
        'procedures': procedures,
        'all_visits': all_visits,
    })


@login_required_custom
def doctor_update_visit(request, doctor_id, visit_id):
    """
    Doctor saves patient data and visit notes.
    DOES NOT change visit status or set completed_at.
    """
    doctor = get_object_or_404(Worker, pk=doctor_id)
    visit = get_object_or_404(Visit, pk=visit_id, doctor=doctor)

    if request.method == 'POST':
        patient = visit.patient
        patient.address = request.POST.get('address', patient.address)
        patient.gender = request.POST.get('gender', patient.gender)
        patient.age = request.POST.get('age') or None
        bd = request.POST.get('birth_date')
        if bd:
            patient.birth_date = bd
        patient.save()

        # Expanded medical notes
        visit.tashxis = request.POST.get('tashxis', visit.tashxis)
        visit.tavsiya = request.POST.get('tavsiya', visit.tavsiya)
        visit.anamnesis_morbi = request.POST.get('anamnesis_morbi', visit.anamnesis_morbi)
        visit.operatsiyalar = request.POST.get('operatsiyalar', visit.operatsiyalar)
        visit.tomizilgan = request.POST.get('tomizilgan', visit.tomizilgan)
        visit.koz_oynak_kontakt_linza = request.POST.get('koz_oynak_kontakt_linza', visit.koz_oynak_kontakt_linza)
        visit.anamnesis_vitae = request.POST.get('anamnesis_vitae', visit.anamnesis_vitae)
        visit.allergiya = request.POST.get('allergiya', visit.allergiya)
        visit.qovoqlar_koz_yosh_yollari = request.POST.get('qovoqlar_koz_yosh_yollari', visit.qovoqlar_koz_yosh_yollari)
        visit.koz_soqqasi = request.POST.get('koz_soqqasi', visit.koz_soqqasi)
        visit.koz_olmasi = request.POST.get('koz_olmasi', visit.koz_olmasi)
        visit.sklera = request.POST.get('sklera', visit.sklera)
        visit.shox_parda = request.POST.get('shox_parda', visit.shox_parda)
        visit.old_kamera = request.POST.get('old_kamera', visit.old_kamera)
        visit.rangdor_parda_qorachiq = request.POST.get('rangdor_parda_qorachiq', visit.rangdor_parda_qorachiq)
        visit.gavhar = request.POST.get('gavhar', visit.gavhar)
        visit.korish_otkirligi = request.POST.get('korish_otkirligi', visit.korish_otkirligi)
        visit.shishasimon_tana = request.POST.get('shishasimon_tana', visit.shishasimon_tana)
        visit.oftalmoskopiya = request.POST.get('oftalmoskopiya', visit.oftalmoskopiya)
        # New OD/OS anatomy fields
        for _f in ['qovoqlar_od', 'qovoqlar_os',
                   'koz_soqqasi_od', 'koz_soqqasi_os',
                   'koz_olmasi_od', 'koz_olmasi_os',
                   'sklera_od', 'sklera_os',
                   'shox_parda_od', 'shox_parda_os',
                   'old_kamera_od', 'old_kamera_os',
                   'rangdor_parda_od', 'rangdor_parda_os',
                   'gavhar_od', 'gavhar_os',
                   'korish_otkirligi_od', 'korish_otkirligi_os',
                   'korreksiya_bilan_od', 'korreksiya_bilan_os',
                   'shishasimon_tana_od', 'shishasimon_tana_os',
                   'oftalmoskopiya_od', 'oftalmoskopiya_os',
                   'fundus_od', 'fundus_os', 'ubm_od', 'ubm_os',
                   'oct_od', 'oct_os', 'kp_od', 'kp_os',
                   'b_skan_od', 'b_skan_os', 'periferiya_od', 'periferiya_os',
                   'retsept_uzoq_od', 'retsept_uzoq_os',
                   'retsept_yaqin_od', 'retsept_yaqin_os',
                   'retsept_kontakt_od', 'retsept_kontakt_os']:
            if _f in request.POST:
                setattr(visit, _f, request.POST.get(_f, ''))
        # Legacy fields
        visit.problem = request.POST.get('problem', visit.problem)
        visit.notes = request.POST.get('notes', visit.notes)

        # Referral partner (doctor can also set this)
        ref_id = request.POST.get('referred_by', '')
        if ref_id:
            visit.referred_by_id = int(ref_id)
        elif 'referred_by' in request.POST:
            visit.referred_by = None

        if request.FILES.get('attachment'):
            visit.attachment = request.FILES['attachment']

        # Keep status as-is. Do NOT auto-complete.
        if visit.status == 'waiting':
            visit.status = 'in_progress'
        visit.save()

        # Handle multiple file attachments
        for f in request.FILES.getlist('attachments'):
            VisitAttachment.objects.create(visit=visit, file=f)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True})

    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_ajax_upload_attachment(request, doctor_id, visit_id):
    """AJAX file upload for doctor inside tibbiy yozuv modal."""
    visit = get_object_or_404(Visit, pk=visit_id)
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        files_data = []
        for f in request.FILES.getlist('attachments'):
            att = VisitAttachment.objects.create(visit=visit, file=f, title=title)
            files_data.append({'url': att.file.url, 'name': f.name, 'title': att.title})
        return JsonResponse({'ok': True, 'files': files_data})
    return JsonResponse({'ok': False}, status=400)


@login_required_custom
def doctor_complete_visit(request, doctor_id, visit_id):
    """
    Explicit "Tugatish" button.
    Always marks the visit as 'done' (completed_at is set).
    Even if there are incomplete procedures, it counts as ended.
    The receptionist can still see procedure reminders separately.
    """
    doctor = get_object_or_404(Worker, pk=doctor_id)
    visit = get_object_or_404(Visit, pk=visit_id, doctor=doctor)

    visit.status = 'done'
    visit.completed_at = timezone.now()
    visit.save()
    return redirect('doctor_panel', doctor_id=doctor_id)


@login_required_custom
def doctor_add_procedure(request, doctor_id, visit_id):
    """
    Add a procedure to a visit.
    DOES NOT change visit status or completed_at.
    The visit stays 'waiting'/'in_progress' until doctor explicitly presses Tugatish.
    """
    doctor = get_object_or_404(Worker, pk=doctor_id)
    visit = get_object_or_404(Visit, pk=visit_id, doctor=doctor)

    if request.method == 'POST':
        description = request.POST.get('procedure_description', '')
        reps_raw = request.POST.get('repetitions', '').strip()
        if not reps_raw or not reps_raw.isdigit() or int(reps_raw) < 1:
            return redirect('doctor_patient_detail', doctor_id=doctor.pk, patient_id=visit.patient.pk)
        repetitions = int(reps_raw)
        price_per_session = int(request.POST.get('price_per_session', 0) or 0)

        procedure = Procedure.objects.create(
            visit=visit,
            patient=visit.patient,
            doctor=doctor,
            description=description,
            total_repetitions=repetitions,
            price_per_session=price_per_session,
            total_price=price_per_session * repetitions,
        )

        for i in range(1, repetitions + 1):
            date_str = request.POST.get(f'date_{i}')
            if date_str:
                ProcedureDate.objects.create(
                    procedure=procedure,
                    scheduled_date=date_str,
                )

        # DO NOT change visit.status here. DO NOT set completed_at.

    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient.id)


@login_required_custom
def doctor_edit_procedure(request, doctor_id, procedure_id):
    doctor = get_object_or_404(Worker, pk=doctor_id)
    procedure = get_object_or_404(Procedure, pk=procedure_id, doctor=doctor)

    if request.method == 'POST':
        procedure.description = request.POST.get('procedure_description', procedure.description)
        reps_raw = request.POST.get('repetitions', '').strip()
        if not reps_raw or not reps_raw.isdigit() or int(reps_raw) < 1:
            return redirect('doctor_patient_detail', doctor_id=doctor.pk, patient_id=procedure.visit.patient.pk)
        new_reps = int(reps_raw)
        price_per_session = int(request.POST.get('price_per_session', procedure.price_per_session) or 0)
        procedure.total_repetitions = new_reps
        procedure.price_per_session = price_per_session
        procedure.total_price = price_per_session * new_reps
        procedure.save()

        procedure.dates.filter(completed=False).delete()

        for i in range(1, new_reps + 1):
            date_str = request.POST.get(f'date_{i}')
            if date_str:
                ProcedureDate.objects.create(
                    procedure=procedure,
                    scheduled_date=date_str,
                )

    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=procedure.visit.patient_id)


@login_required_custom
def doctor_complete_procedure_date(request, doctor_id, pd_id):
    """
    Mark a single procedure date as completed.
    Does NOT auto-complete the visit. The visit's status is managed
    separately via the Tugatish button.
    """
    doctor = get_object_or_404(Worker, pk=doctor_id)
    pd = get_object_or_404(ProcedureDate, pk=pd_id, procedure__doctor=doctor)
    pd.completed = True
    pd.save()

    # Check if ALL procedure dates for this visit are now complete
    procedure = pd.procedure
    visit = procedure.visit

    all_visit_procedures_complete = not ProcedureDate.objects.filter(
        procedure__visit=visit, completed=False
    ).exists()

    if all_visit_procedures_complete and visit.status == 'has_procedure':
        # All procedures done - move visit to 'done'
        visit.status = 'done'
        visit.completed_at = timezone.now()
        visit.save()

    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=procedure.visit.patient_id)


@login_required_custom
def doctor_edit_done_visit(request, doctor_id, visit_id):
    doctor = get_object_or_404(Worker, pk=doctor_id)
    visit = get_object_or_404(Visit, pk=visit_id)

    if request.method == 'POST':
        # Update all medical note fields
        for field in ['tashxis', 'tavsiya', 'anamnesis_morbi',
                      'operatsiyalar', 'tomizilgan', 'koz_oynak_kontakt_linza',
                      'anamnesis_vitae', 'allergiya',
                      'qovoqlar_od', 'qovoqlar_os',
                      'koz_soqqasi_od', 'koz_soqqasi_os',
                      'koz_olmasi_od', 'koz_olmasi_os',
                      'sklera_od', 'sklera_os',
                      'shox_parda_od', 'shox_parda_os',
                      'old_kamera_od', 'old_kamera_os',
                      'rangdor_parda_od', 'rangdor_parda_os',
                      'gavhar_od', 'gavhar_os',
                      'korish_otkirligi_od', 'korish_otkirligi_os',
                      'korreksiya_bilan_od', 'korreksiya_bilan_os',
                      'shishasimon_tana_od', 'shishasimon_tana_os',
                      'fundus_od', 'fundus_os', 'ubm_od', 'ubm_os',
                      'oct_od', 'oct_os', 'kp_od', 'kp_os',
                      'b_skan_od', 'b_skan_os', 'periferiya_od', 'periferiya_os',
                      'oftalmoskopiya_od', 'oftalmoskopiya_os',
                      'retsept_uzoq_od', 'retsept_uzoq_os',
                      'retsept_yaqin_od', 'retsept_yaqin_os',
                      'retsept_kontakt_od', 'retsept_kontakt_os']:
            if field in request.POST:
                setattr(visit, field, request.POST.get(field, ''))
        # Handle file attachments
        if request.FILES.getlist('attachments'):
            for f in request.FILES.getlist('attachments'):
                VisitAttachment.objects.create(visit=visit, file=f)
        visit.save()
    # Redirect back to the referring page
    referer = request.META.get('HTTP_REFERER', '')
    if 'full-history' in referer:
        return redirect('doctor_patient_full_history', doctor_id=doctor_id, patient_id=visit.patient_id)
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
@require_POST
def doctor_delete_attachment(request, doctor_id, att_id):
    """Delete a file attachment from a visit."""
    att = get_object_or_404(VisitAttachment, pk=att_id)
    patient_id = att.visit.patient_id
    att.file.delete(save=False)
    att.delete()
    referer = request.META.get('HTTP_REFERER', '')
    if 'full-history' in referer:
        return redirect('doctor_patient_full_history', doctor_id=doctor_id, patient_id=patient_id)
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=patient_id)


# ─── RECEPTIONIST HISTORY ───

@login_required_custom
def receptionist_history(request):
    user = get_current_user(request)
    if user.role not in ('receptionist', 'admin'):
        return redirect('login')

    # Doctors (active)
    doctors = Worker.objects.filter(worker_type='doctor').order_by('name')
    doctor_list = []
    for d in doctors:
        visit_count = Visit.objects.filter(doctor=d).count()
        patient_count = Visit.objects.filter(doctor=d).values('patient').distinct().count()
        doctor_list.append({
            'doctor': d,
            'visit_count': visit_count,
            'patient_count': patient_count,
        })

    # Deleted doctors — group by doctor_name
    orphan_doctors = {}
    for v in Visit.objects.filter(doctor__isnull=True):
        name = v.doctor_name or "Noma'lum doktor"
        if name not in orphan_doctors:
            orphan_doctors[name] = {'doctor_name': name, 'visit_count': 0}
        orphan_doctors[name]['visit_count'] += 1
    orphan_doctor_list = list(orphan_doctors.values())

    notif_count, proc_count = _get_receptionist_badge_counts()

    return render(request, 'receptionist/history.html', {
        'doctor_list': doctor_list,
        'orphan_doctor_list': orphan_doctor_list,
        'notif_count': notif_count,
        'proc_count': proc_count,
    })


@login_required_custom
def receptionist_patient_history(request, patient_id):
    user = get_current_user(request)
    if user.role not in ('receptionist', 'admin'):
        return redirect('login')

    patient = get_object_or_404(Patient, pk=patient_id)
    visits = Visit.objects.filter(patient=patient).select_related(
        'doctor', 'service'
    ).prefetch_related('procedures__dates').order_by('-created_at')

    total_spent = sum(v.final_price for v in visits.filter(status__in=['done', 'has_procedure']))

    notif_count, proc_count = _get_receptionist_badge_counts()

    return render(request, 'receptionist/patient_history.html', {
        'patient': patient,
        'visits': visits,
        'total_spent': total_spent,
        'notif_count': notif_count,
        'proc_count': proc_count,
    })


@login_required_custom
def receptionist_doctor_history(request, doctor_id):
    user = get_current_user(request)
    if user.role not in ('receptionist', 'admin'):
        return redirect('login')

    doctor = get_object_or_404(Worker, pk=doctor_id)
    visits = Visit.objects.filter(doctor=doctor).select_related(
        'patient', 'service'
    ).order_by('-created_at')

    total_revenue = sum(v.final_price for v in visits.filter(status__in=['done', 'has_procedure']))
    patient_count = visits.values('patient').distinct().count()

    notif_count, proc_count = _get_receptionist_badge_counts()

    return render(request, 'receptionist/doctor_history.html', {
        'doctor': doctor,
        'visits': visits,
        'total_revenue': total_revenue,
        'patient_count': patient_count,
        'notif_count': notif_count,
        'proc_count': proc_count,
    })


@login_required_custom
def receptionist_orphan_visits(request):
    """Show visits where doctor was deleted, allow reassignment."""
    user = get_current_user(request)
    if user.role not in ('receptionist', 'admin'):
        return redirect('login')

    doctor_name_filter = request.GET.get('doctor_name', '')
    visits = Visit.objects.filter(doctor__isnull=True)
    if doctor_name_filter:
        visits = visits.filter(doctor_name=doctor_name_filter)
    visits = visits.select_related('patient', 'service').order_by('-created_at')
    doctors = Worker.objects.filter(worker_type='doctor').order_by('name')

    notif_count, proc_count = _get_receptionist_badge_counts()

    return render(request, 'receptionist/orphan_visits.html', {
        'doctor_name_filter': doctor_name_filter,
        'visits': visits,
        'doctors': doctors,
        'notif_count': notif_count,
        'proc_count': proc_count,
    })


@login_required_custom
def receptionist_reassign_doctor(request):
    """POST: reassign visits from deleted doctor to a new doctor."""
    user = get_current_user(request)
    if user.role not in ('receptionist', 'admin'):
        return redirect('login')

    if request.method == 'POST':
        new_doctor_id = request.POST.get('new_doctor')
        visit_ids = request.POST.getlist('visit_ids')
        if new_doctor_id and visit_ids:
            new_doctor = get_object_or_404(Worker, pk=new_doctor_id, worker_type='doctor')
            Visit.objects.filter(pk__in=visit_ids, doctor__isnull=True).update(
                doctor=new_doctor, doctor_name=new_doctor.full_name
            )
            # Also update procedures
            Procedure.objects.filter(visit_id__in=visit_ids, doctor__isnull=True).update(doctor=new_doctor)
    return redirect('receptionist_orphan_visits')


# ─── API ENDPOINTS FOR CHARTS ───

def api_analytics_overall(request):
    period = request.GET.get('period', 'all')
    date_from = _get_date_filter(period)

    visits = Visit.objects.filter(status__in=['done', 'has_procedure'])
    if date_from:
        visits = visits.filter(created_at__gte=date_from)

    revenue = sum(v.final_price for v in visits)
    doctors_cost = Worker.objects.filter(worker_type='doctor').aggregate(t=Sum('salary'))['t'] or 0
    others_cost = Worker.objects.filter(worker_type='other').aggregate(t=Sum('salary'))['t'] or 0
    monthly_salary = doctors_cost + others_cost

    trends = _get_monthly_trends(period)

    return JsonResponse({
        'revenue': revenue,
        'doctors_cost': doctors_cost,
        'others_cost': others_cost,
        'net': revenue - monthly_salary * _get_period_months(period),
        'trends': trends,
    })


def api_analytics_services(request):
    period = request.GET.get('period', 'all')
    date_from = _get_date_filter(period)

    data = []
    for s in Service.objects.all():
        visits = Visit.objects.filter(
            visit_services__service=s, status__in=['done', 'has_procedure']
        ).distinct()
        if date_from:
            visits = visits.filter(created_at__gte=date_from)
        rev = VisitService.objects.filter(
            service=s,
            visit__status__in=['done', 'has_procedure'],
            **(dict(visit__created_at__gte=date_from) if date_from else {})
        ).aggregate(total=Sum('price_at_time'))['total'] or 0
        data.append({
            'name': s.name,
            'count': visits.count(),
            'revenue': rev,
        })
    return JsonResponse({'data': data})


def api_analytics_doctors(request):
    period = request.GET.get('period', 'all')
    date_from = _get_date_filter(period)

    data = []
    for d in Worker.objects.filter(worker_type='doctor'):
        visits = Visit.objects.filter(doctor=d, status__in=['done', 'has_procedure'])
        if date_from:
            visits = visits.filter(created_at__gte=date_from)
        rated = visits.filter(rated=True)
        avg_r = rated.aggregate(avg=Avg('doctor_rating'))['avg']
        data.append({
            'name': d.full_name,
            'patients': visits.count(),
            'revenue': sum(v.final_price for v in visits),
            'rating': round(avg_r) if avg_r else None,
        })
    return JsonResponse({'data': data})


# ─── PATIENT SEARCH API ───

@login_required_custom
def api_patient_search(request):
    """Live autocomplete search for patients."""
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})

    search_terms = q.split()
    if len(search_terms) >= 2:
        patients = Patient.objects.filter(
            Q(name__icontains=search_terms[0], surname__icontains=search_terms[1]) |
            Q(name__icontains=search_terms[1], surname__icontains=search_terms[0]) |
            Q(phone__icontains=q)
        )
    else:
        patients = Patient.objects.filter(
            Q(phone__icontains=q) |
            Q(name__icontains=q) |
            Q(surname__icontains=q)
        )

    results = [
        {'id': p.pk, 'name': p.full_name, 'phone': p.phone}
        for p in patients[:10]
    ]
    return JsonResponse({'results': results})


# ─── RECEPTIONIST VISIT EDIT ───

@login_required_custom
def receptionist_edit_visit(request, visit_id):
    """Edit a waiting visit (change doctor/services/discount)."""
    user = get_current_user(request)
    if user.role not in ('receptionist', 'admin'):
        return redirect('login')

    visit = get_object_or_404(Visit, pk=visit_id, status='waiting')

    if request.method == 'POST':
        new_doctor_id = request.POST.get('doctor')
        if new_doctor_id:
            new_doctor = get_object_or_404(Worker, pk=new_doctor_id, worker_type='doctor')
            visit.doctor = new_doctor

        # Update services
        visit.visit_services.all().delete()
        service_ids = request.POST.getlist('services')
        if service_ids:
            for sid in service_ids:
                svc = Service.objects.get(pk=sid)
                VisitService.objects.create(visit=visit, service=svc, price_at_time=svc.price)
        elif visit.doctor:
            first_svc = visit.doctor.services.first()
            if first_svc:
                VisitService.objects.create(visit=visit, service=first_svc, price_at_time=first_svc.price)

        discount = int(request.POST.get('discount_amount', '0') or '0')
        visit.discount_amount = discount
        visit.discount_reason = request.POST.get('discount_reason', '')
        visit.save()

    return redirect('receptionist_panel')


# ─── RECEPTIONIST PRINT TICKET ───

@login_required_custom
def receptionist_print_ticket(request, visit_id):
    """Print queue ticket after directing patient to doctor."""
    visit = get_object_or_404(Visit, pk=visit_id)
    queue = getattr(visit, 'queue', None)
    services = visit.visit_services.all()
    procedures = visit.procedures.all()
    return render(request, 'receptionist/ticket.html', {
        'visit': visit,
        'queue': queue,
        'services': services,
        'procedures': procedures,
    })


@login_required_custom
def receptionist_print_lab_ticket(request, lab_visit_id):
    """Print ticket for laboratory visit."""
    lab_visit = get_object_or_404(LabVisit.objects.select_related('patient').prefetch_related('lab_services'), pk=lab_visit_id)
    return render(request, 'receptionist/lab_ticket.html', {
        'lab_visit': lab_visit,
    })


# ─── DOCTOR VISIT RECEIPT ───

@login_required_custom
def doctor_visit_receipt(request, doctor_id, visit_id):
    """Printable receipt after doctor saves."""
    doctor = get_object_or_404(Worker, pk=doctor_id, worker_type='doctor')
    visit = get_object_or_404(Visit, pk=visit_id)
    queue = getattr(visit, 'queue', None)
    services = visit.visit_services.all()
    return render(request, 'doctor/receipt.html', {
        'doctor': doctor,
        'visit': visit,
        'patient': visit.patient,
        'queue': queue,
        'services': services,
    })


@login_required_custom
def doctor_medical_notes(request, doctor_id, visit_id):
    """Printable medical notes (problem + solution + notes) after doctor saves."""
    doctor = get_object_or_404(Worker, pk=doctor_id, worker_type='doctor')
    visit = get_object_or_404(Visit, pk=visit_id)
    return render(request, 'doctor/medical_notes.html', {
        'doctor': doctor,
        'visit': visit,
        'patient': visit.patient,
    })


# ─── SCHEDULING ───

@login_required_custom
def receptionist_schedule(request):
    user = get_current_user(request)
    if user.role not in ('receptionist', 'admin'):
        return redirect('login')

    today = timezone.localdate()
    upcoming = ScheduledVisit.objects.filter(
        scheduled_date__gte=today, activated=False
    ).select_related('patient', 'doctor')

    today_scheduled = upcoming.filter(scheduled_date=today)
    future_scheduled = upcoming.filter(scheduled_date__gt=today)

    doctors = Worker.objects.filter(worker_type='doctor')
    patients = Patient.objects.all().order_by('name')
    notif_count, proc_count = _get_receptionist_badge_counts()

    return render(request, 'receptionist/schedule.html', {
        'today_scheduled': today_scheduled,
        'future_scheduled': future_scheduled,
        'doctors': doctors,
        'patients': patients,
        'notif_count': notif_count,
        'proc_count': proc_count,
    })


@login_required_custom
def add_scheduled_visit(request):
    if request.method == 'POST':
        patient_id = request.POST.get('patient')
        doctor_id = request.POST.get('doctor')
        date = request.POST.get('scheduled_date')
        time_str = request.POST.get('scheduled_time') or None
        notes = request.POST.get('notes', '')
        ScheduledVisit.objects.create(
            patient_id=patient_id,
            doctor_id=doctor_id,
            scheduled_date=date,
            scheduled_time=time_str,
            notes=notes,
        )
    return redirect('receptionist_panel')


@login_required_custom
def activate_scheduled_visit(request, pk):
    """Convert scheduled visit to actual waiting visit (doctor or lab)."""
    sv = get_object_or_404(ScheduledVisit, pk=pk, activated=False)

    # Handle lab scheduled visits
    if sv.is_lab:
        lab_visit = LabVisit.objects.create(patient=sv.patient, status='waiting')
        # Restore lab services from stored data
        if sv.lab_services_data:
            try:
                lab_service_ids = [int(x) for x in json.loads(sv.lab_services_data)]
                svc_map = {t.number: t for t in LabServiceTemplate.objects.filter(number__in=lab_service_ids)}
                for ls_id in lab_service_ids:
                    tpl = svc_map.get(ls_id)
                    if tpl:
                        LabVisitService.objects.create(
                            lab_visit=lab_visit,
                            service_number=tpl.number,
                            service_name=tpl.name,
                            price=tpl.price,
                        )
            except (json.JSONDecodeError, ValueError):
                pass
        # Lab queue number
        try:
            lqn = get_next_lab_queue_number()
            LabQueueNumber.objects.create(lab_visit=lab_visit, number=lqn, date=timezone.localdate())
        except Exception:
            pass
        sv.activated = True
        sv.save()
        return redirect('receptionist_print_lab_ticket', lab_visit_id=lab_visit.pk)

    doctor = sv.doctor

    first_svc = doctor.services.first() if doctor else None
    visit = Visit.objects.create(
        patient=sv.patient,
        doctor=doctor,
        doctor_name=doctor.full_name if doctor else '',
        service=first_svc,
        discount_amount=0,
        status='waiting',
    )
    # Use scheduled services if available, otherwise doctor's first service
    sv_services = sv.services.all()
    if sv_services.exists():
        for svc in sv_services:
            VisitService.objects.create(visit=visit, service=svc, price_at_time=svc.price)
    elif first_svc:
        VisitService.objects.create(visit=visit, service=first_svc, price_at_time=first_svc.price)

    qnum = get_next_queue_number(doctor=doctor)
    QueueNumber.objects.create(visit=visit, doctor=doctor, number=qnum, date=timezone.localdate())

    sv.activated = True
    sv.visit = visit
    sv.save()

    return redirect('receptionist_print_ticket', visit_id=visit.pk)


@login_required_custom
def edit_scheduled_visit(request, pk):
    sv = get_object_or_404(ScheduledVisit, pk=pk, activated=False)
    if request.method == 'POST':
        doctor_id = request.POST.get('doctor')
        if doctor_id:
            sv.doctor = get_object_or_404(Worker, pk=doctor_id, worker_type='doctor')
        date_str = request.POST.get('scheduled_date', '')
        if date_str:
            try:
                sv.scheduled_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        sv.scheduled_time = request.POST.get('scheduled_time') or None
        sv.notes = request.POST.get('notes', sv.notes)
        sv.save()
        # Update services M2M
        service_ids = request.POST.getlist('services')
        sv.services.set(service_ids)
    return redirect('receptionist_panel')


@login_required_custom
def delete_scheduled_visit(request, pk):
    sv = get_object_or_404(ScheduledVisit, pk=pk)
    sv.delete()
    return redirect('receptionist_panel')


# ─── API: Patient Procedures ───

@login_required_custom
def api_patient_procedures(request, patient_id):
    """Get unique procedures for a patient (for receptionist redirect dropdown)."""
    patient = get_object_or_404(Patient, pk=patient_id)
    procedures = Procedure.objects.filter(patient=patient).select_related('doctor')
    results = []
    seen = set()
    for p in procedures:
        key = p.description.strip().lower()
        if key not in seen:
            seen.add(key)
            results.append({
                'id': p.pk,
                'description': p.description,
                'price_per_session': p.price_per_session,
                'total_price': p.total_price,
                'doctor_name': p.doctor.full_name if p.doctor else '',
            })
    return JsonResponse({'procedures': results})


# ─── DATA EXPORT ───

@role_required('admin')
def export_data_excel(request):
    import openpyxl
    from django.http import HttpResponse as HR

    wb = openpyxl.Workbook()

    # Single sheet: Bemorlar tashriflari (matching the table on analytics page)
    ws = wb.active
    ws.title = "Bemorlar tashriflari"
    ws.append(['#', 'Bemor', 'Telefon', 'Doktor', 'Xizmat', 'Sana', 'Chegirma'])

    visits = Visit.objects.select_related('patient', 'doctor', 'service').prefetch_related(
        'visit_services__service'
    ).order_by('-created_at')

    for i, v in enumerate(visits, 1):
        svc_names = ', '.join(vs.service.name for vs in v.visit_services.all() if vs.service)
        if not svc_names and v.service:
            svc_names = v.service.name
        discount_text = ''
        if v.discount_amount:
            discount_text = f"-{v.discount_amount:,} so'm".replace(',', ' ')
            if v.discount_reason:
                discount_text += f' ({v.discount_reason})'
        ws.append([
            i,
            v.patient.full_name,
            v.patient.phone,
            v.doctor.full_name if v.doctor else (v.doctor_name or ''),
            svc_names,
            v.created_at.strftime('%d.%m.%Y'),
            discount_text or '—',
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="optimed_data_export.xlsx"'
    wb.save(response)
    return response


@role_required('admin')
def export_data_pdf(request):
    """Export clinic data as a styled Excel (PDF alternative) with visit table."""
    from django.http import HttpResponse as HR
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def style_data(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border

    visits = Visit.objects.select_related('patient', 'doctor', 'service').prefetch_related(
        'visit_services__service'
    ).order_by('-created_at')

    ws = wb.active
    ws.title = "Bemorlar tashriflari"
    headers = ['#', 'Bemor', 'Telefon', 'Doktor', 'Xizmat', 'Sana', 'Chegirma']
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for i, v in enumerate(visits, 1):
        svc_names = ', '.join(vs.service.name for vs in v.visit_services.all() if vs.service)
        if not svc_names and v.service:
            svc_names = v.service.name
        discount_text = ''
        if v.discount_amount:
            discount_text = f"-{v.discount_amount:,} so'm".replace(',', ' ')
            if v.discount_reason:
                discount_text += f' ({v.discount_reason})'
        ws.append([
            i,
            v.patient.full_name,
            v.patient.phone,
            v.doctor.full_name if v.doctor else (v.doctor_name or ''),
            svc_names,
            v.created_at.strftime('%d.%m.%Y'),
            discount_text or '—',
        ])
        style_data(ws, i + 1, len(headers))

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="optimed_hisobot_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# ─── GLASSES MANAGEMENT (ADMIN) ───

@role_required('admin')
def add_glasses(request):
    if request.method == 'POST':
        Glasses.objects.create(
            model_name=request.POST['model_name'],
            cost_price=int(request.POST.get('cost_price', 0) or 0),
            sell_price=int(request.POST.get('sell_price', 0) or 0),
            lens_type=request.POST.get('lens_type', ''),
            brand=request.POST.get('brand', ''),
            quantity=int(request.POST.get('quantity', 0) or 0),
        )
    return redirect('admin_panel')


@role_required('admin')
def edit_glasses(request, pk):
    g = get_object_or_404(Glasses, pk=pk)
    if request.method == 'POST':
        g.model_name = request.POST['model_name']
        g.cost_price = int(request.POST.get('cost_price', 0) or 0)
        g.sell_price = int(request.POST.get('sell_price', 0) or 0)
        g.lens_type = request.POST.get('lens_type', '')
        g.brand = request.POST.get('brand', '')
        g.quantity = int(request.POST.get('quantity', 0) or 0)
        g.save()
    return redirect('admin_panel')


@role_required('admin')
def delete_glasses(request, pk):
    get_object_or_404(Glasses, pk=pk).delete()
    return redirect('admin_panel')


# ─── SELLER PANEL ───

@login_required_custom
@role_required('seller')
def seller_panel(request):
    glasses = Glasses.objects.all()
    return render(request, 'seller/panel.html', {'glasses': glasses})


@role_required('seller')
def seller_add_glasses(request):
    if request.method == 'POST':
        Glasses.objects.create(
            model_name=request.POST['model_name'],
            cost_price=int(request.POST.get('cost_price', 0) or 0),
            sell_price=int(request.POST.get('sell_price', 0) or 0),
            lens_type=request.POST.get('lens_type', ''),
            brand=request.POST.get('brand', ''),
            quantity=int(request.POST.get('quantity', 0) or 0),
        )
    return redirect('seller_panel')


@role_required('seller')
def seller_edit_glasses(request, pk):
    g = get_object_or_404(Glasses, pk=pk)
    if request.method == 'POST':
        g.model_name = request.POST['model_name']
        g.cost_price = int(request.POST.get('cost_price', 0) or 0)
        g.sell_price = int(request.POST.get('sell_price', 0) or 0)
        g.lens_type = request.POST.get('lens_type', '')
        g.brand = request.POST.get('brand', '')
        g.quantity = int(request.POST.get('quantity', 0) or 0)
        g.save()
    return redirect('seller_panel')


@role_required('seller')
def seller_delete_glasses(request, pk):
    get_object_or_404(Glasses, pk=pk).delete()
    return redirect('seller_panel')


@role_required('seller')
def seller_sell_glasses(request):
    if request.method == 'POST':
        glasses_id = request.POST.get('glasses_id')
        discount = int(request.POST.get('discount_amount', 0) or 0)
        g = get_object_or_404(Glasses, pk=glasses_id)
        if g.quantity < 1:
            return redirect('seller_sales')
        GlassSale.objects.create(
            glasses=g,
            quantity=1,
            sell_price=g.sell_price,
            cost_price=g.cost_price,
            discount_amount=discount,
        )
        g.quantity -= 1
        g.save(update_fields=['quantity'])
    return redirect('seller_sales')


@role_required('seller')
def seller_sales(request):
    glasses = Glasses.objects.filter(quantity__gt=0).order_by('model_name')
    return render(request, 'seller/sales.html', {'glasses': glasses})


@role_required('seller')
def seller_analytics(request):
    period = request.GET.get('period', 'month')
    date_from = _get_date_filter(period)

    sales_qs = GlassSale.objects.all()
    if date_from:
        sales_qs = sales_qs.filter(created_at__gte=date_from)

    total_revenue = sum(s.total_revenue for s in sales_qs)
    total_profit = sum(s.total_profit for s in sales_qs)
    total_sold = sales_qs.aggregate(total=Sum('quantity'))['total'] or 0

    return render(request, 'seller/analytics.html', {
        'period': period,
        'total_revenue': total_revenue,
        'total_profit': total_profit,
        'total_sold': total_sold,
        'sales': sales_qs.select_related('glasses')[:50],
    })


# ─── LABORATORY PANEL ───

from .lab_services import LAB_SERVICES, search_lab_services


@role_required('laboratory')
def lab_panel(request):
    lab_visits = LabVisit.objects.select_related('patient').prefetch_related('lab_services').filter(
        status__in=['waiting', 'in_progress']
    )
    done_visits = LabVisit.objects.select_related('patient').prefetch_related('lab_services').filter(
        status='done'
    ).order_by('-completed_at')[:50]
    return render(request, 'lab/panel.html', {
        'lab_visits': lab_visits,
        'done_visits': done_visits,
    })


@role_required('laboratory')
def lab_visit_detail(request, pk):
    lv = get_object_or_404(LabVisit.objects.select_related('patient').prefetch_related('lab_services'), pk=pk)
    if lv.status == 'waiting':
        lv.status = 'in_progress'
        lv.save(update_fields=['status'])
    # Build reference values map from DB
    ref_map = {t.number: t.reference_value for t in LabServiceTemplate.objects.all()}
    ref_values = {}
    for svc in lv.lab_services.all():
        ref_values[svc.pk] = ref_map.get(svc.service_number, '')
    return render(request, 'lab/visit_detail.html', {'lab_visit': lv, 'ref_values': ref_values})


@role_required('laboratory')
def lab_save_results(request, pk):
    lv = get_object_or_404(LabVisit, pk=pk)
    if request.method == 'POST':
        for ls in lv.lab_services.all():
            result = request.POST.get(f'result_{ls.pk}', '').strip()
            if result != ls.result:
                ls.result = result
                ls.save(update_fields=['result'])
        action = request.POST.get('action', '')
        if action == 'done':
            lv.status = 'done'
            lv.completed_at = timezone.now()
            lv.save(update_fields=['status', 'completed_at'])
            return redirect('lab_panel')
    return redirect('lab_visit_detail', pk=pk)


# ─── API: Lab services search ───

def api_lab_services_search(request):
    q = request.GET.get('q', '').strip()
    num = request.GET.get('num', '').strip()
    if len(q) < 1 and len(num) < 1:
        return JsonResponse({'results': []})
    qs = LabServiceTemplate.objects.all()
    if num:
        qs = qs.filter(number__startswith=num)
    if q:
        if q.isdigit() and not num:
            qs = qs.filter(number__startswith=q)
        else:
            qs = qs.filter(name__icontains=q)
    results = qs[:30]
    return JsonResponse({'results': [
        {'number': svc.number, 'name': svc.name, 'price': svc.price}
        for svc in results
    ]})


def api_check_phone(request):
    """Check if a phone number is already registered."""
    phone = normalize_phone(request.GET.get('phone', ''))
    try:
        patient = Patient.objects.get(phone=phone)
        return JsonResponse({'exists': True, 'patient_name': patient.full_name})
    except Patient.DoesNotExist:
        return JsonResponse({'exists': False})


# ─── DOCTOR: Eye examination data (REF, KRT, IOP) ───

@login_required_custom
def doctor_add_refraction(request, doctor_id, visit_id):
    visit = get_object_or_404(Visit, pk=visit_id)
    if request.method == 'POST':
        ref = EyeRefraction.objects.create(
            visit=visit, patient=visit.patient,
            od_s=request.POST.get('od_s', ''),
            od_c=request.POST.get('od_c', ''),
            od_a=request.POST.get('od_a', ''),
            os_s=request.POST.get('os_s', ''),
            os_c=request.POST.get('os_c', ''),
            os_a=request.POST.get('os_a', ''),
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'date': ref.created_at.strftime('%d.%m.%Y'),
                'od_s': ref.od_s, 'od_c': ref.od_c, 'od_a': ref.od_a,
                'os_s': ref.os_s, 'os_c': ref.os_c, 'os_a': ref.os_a})
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_edit_refraction(request, doctor_id, pk):
    ref = get_object_or_404(EyeRefraction, pk=pk)
    if request.method == 'POST':
        ref.od_s = request.POST.get('od_s', '')
        ref.od_c = request.POST.get('od_c', '')
        ref.od_a = request.POST.get('od_a', '')
        ref.os_s = request.POST.get('os_s', '')
        ref.os_c = request.POST.get('os_c', '')
        ref.os_a = request.POST.get('os_a', '')
        ref.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=ref.patient_id)


@login_required_custom
def doctor_delete_refraction(request, doctor_id, pk):
    ref = get_object_or_404(EyeRefraction, pk=pk)
    pid = ref.patient_id
    ref.delete()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=pid)


@login_required_custom
def doctor_add_krt(request, doctor_id, visit_id):
    visit = get_object_or_404(Visit, pk=visit_id)
    if request.method == 'POST':
        krt = EyeKRT.objects.create(
            visit=visit, patient=visit.patient,
            od_d=request.POST.get('od_d', ''),
            od_mm=request.POST.get('od_mm', ''),
            od_a=request.POST.get('od_a', ''),
            os_d=request.POST.get('os_d', ''),
            os_mm=request.POST.get('os_mm', ''),
            os_a=request.POST.get('os_a', ''),
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'date': krt.created_at.strftime('%d.%m.%Y'),
                'od_d': krt.od_d, 'od_mm': krt.od_mm, 'od_a': krt.od_a,
                'os_d': krt.os_d, 'os_mm': krt.os_mm, 'os_a': krt.os_a})
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_edit_krt(request, doctor_id, pk):
    krt = get_object_or_404(EyeKRT, pk=pk)
    if request.method == 'POST':
        krt.od_d = request.POST.get('od_d', '')
        krt.od_mm = request.POST.get('od_mm', '')
        krt.od_a = request.POST.get('od_a', '')
        krt.os_d = request.POST.get('os_d', '')
        krt.os_mm = request.POST.get('os_mm', '')
        krt.os_a = request.POST.get('os_a', '')
        krt.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=krt.patient_id)


@login_required_custom
def doctor_delete_krt(request, doctor_id, pk):
    krt = get_object_or_404(EyeKRT, pk=pk)
    pid = krt.patient_id
    krt.delete()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=pid)


@login_required_custom
def doctor_add_iop(request, doctor_id, visit_id):
    visit = get_object_or_404(Visit, pk=visit_id)
    if request.method == 'POST':
        iop = EyeIOP.objects.create(
            visit=visit, patient=visit.patient,
            od_value=request.POST.get('od_value', ''),
            os_value=request.POST.get('os_value', ''),
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'date': iop.created_at.strftime('%d.%m.%Y'),
                'od_value': iop.od_value, 'os_value': iop.os_value})
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_edit_iop(request, doctor_id, pk):
    iop = get_object_or_404(EyeIOP, pk=pk)
    if request.method == 'POST':
        iop.od_value = request.POST.get('od_value', '')
        iop.os_value = request.POST.get('os_value', '')
        iop.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=iop.patient_id)


@login_required_custom
def doctor_delete_iop(request, doctor_id, pk):
    iop = get_object_or_404(EyeIOP, pk=pk)
    pid = iop.patient_id
    iop.delete()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=pid)


# ─── VISUAL ACUITY CRUD ───

@login_required_custom
def doctor_add_visual_acuity(request, doctor_id, visit_id):
    visit = get_object_or_404(Visit, pk=visit_id)
    if request.method == 'POST':
        va = EyeVisualAcuity.objects.create(
            visit=visit, patient=visit.patient,
            uzoq_od_sph=request.POST.get('uzoq_od_sph', ''),
            uzoq_od_cyl=request.POST.get('uzoq_od_cyl', ''),
            uzoq_od_ax=request.POST.get('uzoq_od_ax', ''),
            uzoq_od_vis=request.POST.get('uzoq_od_vis', ''),
            uzoq_os_sph=request.POST.get('uzoq_os_sph', ''),
            uzoq_os_cyl=request.POST.get('uzoq_os_cyl', ''),
            uzoq_os_ax=request.POST.get('uzoq_os_ax', ''),
            uzoq_os_vis=request.POST.get('uzoq_os_vis', ''),
            yaqin_od_sph=request.POST.get('yaqin_od_sph', ''),
            yaqin_od_cyl=request.POST.get('yaqin_od_cyl', ''),
            yaqin_od_ax=request.POST.get('yaqin_od_ax', ''),
            yaqin_od_vis=request.POST.get('yaqin_od_vis', ''),
            yaqin_os_sph=request.POST.get('yaqin_os_sph', ''),
            yaqin_os_cyl=request.POST.get('yaqin_os_cyl', ''),
            yaqin_os_ax=request.POST.get('yaqin_os_ax', ''),
            yaqin_os_vis=request.POST.get('yaqin_os_vis', ''),
            mkl_od_sph=request.POST.get('mkl_od_sph', ''),
            mkl_od_cyl=request.POST.get('mkl_od_cyl', ''),
            mkl_od_ax=request.POST.get('mkl_od_ax', ''),
            mkl_od_vis=request.POST.get('mkl_od_vis', ''),
            mkl_os_sph=request.POST.get('mkl_os_sph', ''),
            mkl_os_cyl=request.POST.get('mkl_os_cyl', ''),
            mkl_os_ax=request.POST.get('mkl_os_ax', ''),
            mkl_os_vis=request.POST.get('mkl_os_vis', ''),
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'date': va.created_at.strftime('%d.%m.%Y'),
                'uzoq_od_sph': va.uzoq_od_sph, 'uzoq_od_cyl': va.uzoq_od_cyl,
                'uzoq_od_ax': va.uzoq_od_ax, 'uzoq_od_vis': va.uzoq_od_vis,
                'uzoq_os_sph': va.uzoq_os_sph, 'uzoq_os_cyl': va.uzoq_os_cyl,
                'uzoq_os_ax': va.uzoq_os_ax, 'uzoq_os_vis': va.uzoq_os_vis,
                'yaqin_od_sph': va.yaqin_od_sph, 'yaqin_od_vis': va.yaqin_od_vis,
                'yaqin_os_sph': va.yaqin_os_sph, 'yaqin_os_vis': va.yaqin_os_vis,
                'mkl_od_sph': va.mkl_od_sph, 'mkl_od_vis': va.mkl_od_vis,
                'mkl_os_sph': va.mkl_os_sph, 'mkl_os_vis': va.mkl_os_vis})
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_edit_visual_acuity(request, doctor_id, pk):
    va = get_object_or_404(EyeVisualAcuity, pk=pk)
    if request.method == 'POST':
        for field in ['uzoq_od_sph', 'uzoq_od_cyl', 'uzoq_od_ax', 'uzoq_od_vis',
                       'uzoq_os_sph', 'uzoq_os_cyl', 'uzoq_os_ax', 'uzoq_os_vis',
                       'yaqin_od_sph', 'yaqin_od_cyl', 'yaqin_od_ax', 'yaqin_od_vis',
                       'yaqin_os_sph', 'yaqin_os_cyl', 'yaqin_os_ax', 'yaqin_os_vis',
                       'mkl_od_sph', 'mkl_od_cyl', 'mkl_od_ax', 'mkl_od_vis',
                       'mkl_os_sph', 'mkl_os_cyl', 'mkl_os_ax', 'mkl_os_vis']:
            setattr(va, field, request.POST.get(field, ''))
        va.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=va.patient_id)


@login_required_custom
def doctor_delete_visual_acuity(request, doctor_id, pk):
    va = get_object_or_404(EyeVisualAcuity, pk=pk)
    patient_id = va.patient_id
    va.delete()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=patient_id)


# ─── EYE EXAM FORM CRUD ───

@login_required_custom
def doctor_add_eye_exam(request, doctor_id, visit_id):
    visit = get_object_or_404(Visit, pk=visit_id)
    doctor = get_object_or_404(Worker, pk=doctor_id)
    if request.method == 'POST':
        exam = EyeExamForm(
            patient=visit.patient, visit=visit, doctor=doctor,
            protocol_number=request.POST.get('protocol_number', ''),
            kelish_sanasi=timezone.now(), filial='Optimed',
        )
        for field in ['koz_anamezi', 'oyku',
                      'ref_ong_sph', 'ref_ong_cyl', 'ref_ong_axis',
                      'ref_chap_sph', 'ref_chap_cyl', 'ref_chap_axis',
                      'sub_ref_ong_sph', 'sub_ref_ong_cyl', 'sub_ref_ong_axis',
                      'sub_ref_chap_sph', 'sub_ref_chap_cyl', 'sub_ref_chap_axis',
                      'sik_ref_ong_sph', 'sik_ref_ong_cyl', 'sik_ref_ong_axis',
                      'sik_ref_chap_sph', 'sik_ref_chap_cyl', 'sik_ref_chap_axis',
                      'tuzatilmagan_korish_ong', 'tuzatilmagan_korish_chap',
                      'tuzatilgan_korish_ong', 'tuzatilgan_korish_chap',
                      'old_segment_ong', 'old_segment_chap',
                      'orqa_segment_ong', 'orqa_segment_chap',
                      'koz_bosimi_ong', 'koz_bosimi_chap',
                      'uzoq_ong_sph', 'uzoq_ong_cyl', 'uzoq_ong_axis',
                      'uzoq_chap_sph', 'uzoq_chap_cyl', 'uzoq_chap_axis',
                      'yaqin_ong_sph', 'yaqin_ong_cyl', 'yaqin_ong_axis',
                      'yaqin_chap_sph', 'yaqin_chap_cyl', 'yaqin_chap_axis',
                      'linza_ong', 'linza_chap', 'diametr_ong', 'diametr_chap',
                      'radius_ong', 'radius_chap', 'dioptriya_ong', 'dioptriya_chap',
                      'icd_kodi', 'icd_nomi', 'tashxis_turi', 'yonalish']:
            setattr(exam, field, request.POST.get(field, ''))
        exam.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_edit_eye_exam(request, doctor_id, pk):
    exam = get_object_or_404(EyeExamForm, pk=pk)
    if request.method == 'POST':
        for field in ['koz_anamezi', 'oyku',
                      'ref_ong_sph', 'ref_ong_cyl', 'ref_ong_axis',
                      'ref_chap_sph', 'ref_chap_cyl', 'ref_chap_axis',
                      'sub_ref_ong_sph', 'sub_ref_ong_cyl', 'sub_ref_ong_axis',
                      'sub_ref_chap_sph', 'sub_ref_chap_cyl', 'sub_ref_chap_axis',
                      'sik_ref_ong_sph', 'sik_ref_ong_cyl', 'sik_ref_ong_axis',
                      'sik_ref_chap_sph', 'sik_ref_chap_cyl', 'sik_ref_chap_axis',
                      'tuzatilmagan_korish_ong', 'tuzatilmagan_korish_chap',
                      'tuzatilgan_korish_ong', 'tuzatilgan_korish_chap',
                      'old_segment_ong', 'old_segment_chap',
                      'orqa_segment_ong', 'orqa_segment_chap',
                      'koz_bosimi_ong', 'koz_bosimi_chap',
                      'uzoq_ong_sph', 'uzoq_ong_cyl', 'uzoq_ong_axis',
                      'uzoq_chap_sph', 'uzoq_chap_cyl', 'uzoq_chap_axis',
                      'yaqin_ong_sph', 'yaqin_ong_cyl', 'yaqin_ong_axis',
                      'yaqin_chap_sph', 'yaqin_chap_cyl', 'yaqin_chap_axis',
                      'linza_ong', 'linza_chap', 'diametr_ong', 'diametr_chap',
                      'radius_ong', 'radius_chap', 'dioptriya_ong', 'dioptriya_chap',
                      'icd_kodi', 'icd_nomi', 'tashxis_turi', 'yonalish']:
            setattr(exam, field, request.POST.get(field, ''))
        exam.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=exam.patient_id)


@login_required_custom
def doctor_delete_eye_exam(request, doctor_id, pk):
    exam = get_object_or_404(EyeExamForm, pk=pk)
    patient_id = exam.patient_id
    exam.delete()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=patient_id)


# ─── DISCHARGE FORM CRUD ───

@login_required_custom
def doctor_add_discharge(request, doctor_id, visit_id):
    visit = get_object_or_404(Visit, pk=visit_id)
    doctor = get_object_or_404(Worker, pk=doctor_id)
    if request.method == 'POST':
        form = DischargeForm(
            patient=visit.patient, visit=visit, doctor=doctor,
            protocol_number=request.POST.get('protocol_number', ''),
        )
        yotish = request.POST.get('yotish_sanasi', '')
        chiqish = request.POST.get('chiqish_sanasi', '')
        form.yotish_sanasi = yotish if yotish else None
        form.chiqish_sanasi = chiqish if chiqish else None
        for field in ['shikoyat_tarix', 'shaxsiy_oilaviy_tarix', 'tizimli_anamnez',
                      'koz_anamezi', 'tashxis_kodi_nomi', 'operatsiya_kodi_nomi',
                      'old_segment_ong', 'old_segment_chap',
                      'fundus_ong', 'fundus_chap',
                      'preop_ong', 'preop_chap',
                      'postop_ong', 'postop_chap',
                      'koz_bosimi_ong', 'koz_bosimi_chap',
                      'daraja_ong', 'daraja_chap',
                      'qovoq_ong', 'qovoq_chap',
                      'glob_ong', 'glob_chap',
                      'muhim_tekshiruv', 'anesteziya_turi', 'operatsiya_izohlar',
                      'chiqish_holati', 'davolash', 'anesteziya_izohi', 'ishlatilgan_linzalar']:
            setattr(form, field, request.POST.get(field, ''))
        form.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_edit_discharge(request, doctor_id, pk):
    form_obj = get_object_or_404(DischargeForm, pk=pk)
    if request.method == 'POST':
        yotish = request.POST.get('yotish_sanasi', '')
        chiqish = request.POST.get('chiqish_sanasi', '')
        form_obj.yotish_sanasi = yotish if yotish else None
        form_obj.chiqish_sanasi = chiqish if chiqish else None
        for field in ['shikoyat_tarix', 'shaxsiy_oilaviy_tarix', 'tizimli_anamnez',
                      'koz_anamezi', 'tashxis_kodi_nomi', 'operatsiya_kodi_nomi',
                      'old_segment_ong', 'old_segment_chap',
                      'fundus_ong', 'fundus_chap',
                      'preop_ong', 'preop_chap',
                      'postop_ong', 'postop_chap',
                      'koz_bosimi_ong', 'koz_bosimi_chap',
                      'daraja_ong', 'daraja_chap',
                      'qovoq_ong', 'qovoq_chap',
                      'glob_ong', 'glob_chap',
                      'muhim_tekshiruv', 'anesteziya_turi', 'operatsiya_izohlar',
                      'chiqish_holati', 'davolash', 'anesteziya_izohi', 'ishlatilgan_linzalar']:
            setattr(form_obj, field, request.POST.get(field, ''))
        form_obj.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=form_obj.patient_id)


@login_required_custom
def doctor_delete_discharge(request, doctor_id, pk):
    form_obj = get_object_or_404(DischargeForm, pk=pk)
    patient_id = form_obj.patient_id
    form_obj.delete()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=patient_id)


# ─── LASIK FORM CRUD ───

@login_required_custom
def doctor_add_lasik(request, doctor_id, visit_id):
    visit = get_object_or_404(Visit, pk=visit_id)
    doctor = get_object_or_404(Worker, pk=doctor_id)
    if request.method == 'POST':
        obj = LasikForm(
            patient=visit.patient, visit=visit, doctor=doctor,
            protocol_number=request.POST.get('protocol_number', ''),
            octavius=request.POST.get('octavius') == 'on',
        )
        op_sana = request.POST.get('operatsiya_sanasi', '')
        obj.operatsiya_sanasi = op_sana if op_sana else None
        for field in ['tashxis', 'operatsiya_turi', 'ko_z',
                      'oper_hirurg', 'assistent',
                      'vgd_od', 'vgd_os',
                      'preop_vis_od', 'preop_sph_od', 'preop_cyl_od', 'preop_axis_od',
                      'preop_vis_os', 'preop_sph_os', 'preop_cyl_os', 'preop_axis_os',
                      'postop_vis_od', 'postop_vis_os', 'preop_corr_od', 'preop_corr_os',
                      'holat', 'tavsiya']:
            setattr(obj, field, request.POST.get(field, ''))
        obj.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_edit_lasik(request, doctor_id, pk):
    obj = get_object_or_404(LasikForm, pk=pk)
    if request.method == 'POST':
        obj.protocol_number = request.POST.get('protocol_number', obj.protocol_number)
        obj.octavius = request.POST.get('octavius') == 'on'
        op_sana = request.POST.get('operatsiya_sanasi', '')
        obj.operatsiya_sanasi = op_sana if op_sana else None
        for field in ['tashxis', 'operatsiya_turi', 'ko_z',
                      'oper_hirurg', 'assistent',
                      'vgd_od', 'vgd_os',
                      'preop_vis_od', 'preop_sph_od', 'preop_cyl_od', 'preop_axis_od',
                      'preop_vis_os', 'preop_sph_os', 'preop_cyl_os', 'preop_axis_os',
                      'postop_vis_od', 'postop_vis_os', 'preop_corr_od', 'preop_corr_os',
                      'holat', 'tavsiya']:
            setattr(obj, field, request.POST.get(field, ''))
        obj.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=obj.patient_id)


@login_required_custom
def doctor_delete_lasik(request, doctor_id, pk):
    obj = get_object_or_404(LasikForm, pk=pk)
    patient_id = obj.patient_id
    obj.delete()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=patient_id)


# ─── STRABISMUS FORM CRUD ───

@login_required_custom
def doctor_add_strabismus(request, doctor_id, visit_id):
    visit = get_object_or_404(Visit, pk=visit_id)
    doctor = get_object_or_404(Worker, pk=doctor_id)
    if request.method == 'POST':
        obj = StrabismusForm(
            patient=visit.patient, visit=visit, doctor=doctor,
            protocol_number=request.POST.get('protocol_number', ''),
        )
        for field in ['shikoyat', 'vis', 'dev_darajasi', 'tashxis',
                      'recession_muscle', 'recession_eye', 'recession_mm',
                      'resection_muscle', 'resection_eye', 'resection_mm',
                      'transposition_eye', 'postop_dev', 'tavsiya',
                      'gilaylik_keyin', 'gilaylik_keyin_ru', 'dorilar', 'dorilar_ru']:
            setattr(obj, field, request.POST.get(field, ''))
        obj.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_edit_strabismus(request, doctor_id, pk):
    obj = get_object_or_404(StrabismusForm, pk=pk)
    if request.method == 'POST':
        obj.protocol_number = request.POST.get('protocol_number', obj.protocol_number)
        for field in ['shikoyat', 'vis', 'dev_darajasi', 'tashxis',
                      'recession_muscle', 'recession_eye', 'recession_mm',
                      'resection_muscle', 'resection_eye', 'resection_mm',
                      'transposition_eye', 'postop_dev', 'tavsiya',
                      'gilaylik_keyin', 'gilaylik_keyin_ru', 'dorilar', 'dorilar_ru']:
            setattr(obj, field, request.POST.get(field, ''))
        obj.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=obj.patient_id)


@login_required_custom
def doctor_delete_strabismus(request, doctor_id, pk):
    obj = get_object_or_404(StrabismusForm, pk=pk)
    patient_id = obj.patient_id
    obj.delete()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=patient_id)


# ─── POST-OP FORM CRUD ───

@login_required_custom
def doctor_add_postop(request, doctor_id, visit_id):
    visit = get_object_or_404(Visit, pk=visit_id)
    doctor = get_object_or_404(Worker, pk=doctor_id)
    if request.method == 'POST':
        obj = PostOpForm(
            patient=visit.patient, visit=visit, doctor=doctor,
            protocol_number=request.POST.get('protocol_number', ''),
        )
        bosh = request.POST.get('davolash_boshlanishi', '')
        tug = request.POST.get('davolash_tugashi', '')
        op_sana = request.POST.get('operatsiya_sanasi', '')
        obj.davolash_boshlanishi = bosh if bosh else None
        obj.davolash_tugashi = tug if tug else None
        obj.operatsiya_sanasi = op_sana if op_sana else None
        for field in ['asosiy_tashxis', 'qoshimcha_tashxis', 'hamroh_kasallik',
                      'qabulda_vis_od', 'qabulda_vis_os', 'qabulda_od', 'qabulda_os',
                      'xususiyatlar', 'operatsiya', 'narkoz_turi', 'hirurg',
                      'chiqimda_vis_od', 'chiqimda_vis_os', 'chiqimda_od', 'chiqimda_os',
                      'xususiyat_chiqim', 'tavsiya', 'keyingi_qayta_korish', 'lech_vrach',
                      'bemor_eslatma', 'bemor_eslatma_ru']:
            setattr(obj, field, request.POST.get(field, ''))
        obj.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=visit.patient_id)


@login_required_custom
def doctor_edit_postop(request, doctor_id, pk):
    obj = get_object_or_404(PostOpForm, pk=pk)
    if request.method == 'POST':
        obj.protocol_number = request.POST.get('protocol_number', obj.protocol_number)
        bosh = request.POST.get('davolash_boshlanishi', '')
        tug = request.POST.get('davolash_tugashi', '')
        op_sana = request.POST.get('operatsiya_sanasi', '')
        obj.davolash_boshlanishi = bosh if bosh else None
        obj.davolash_tugashi = tug if tug else None
        obj.operatsiya_sanasi = op_sana if op_sana else None
        for field in ['asosiy_tashxis', 'qoshimcha_tashxis', 'hamroh_kasallik',
                      'qabulda_vis_od', 'qabulda_vis_os', 'qabulda_od', 'qabulda_os',
                      'xususiyatlar', 'operatsiya', 'narkoz_turi', 'hirurg',
                      'chiqimda_vis_od', 'chiqimda_vis_os', 'chiqimda_od', 'chiqimda_os',
                      'xususiyat_chiqim', 'tavsiya', 'keyingi_qayta_korish', 'lech_vrach',
                      'bemor_eslatma', 'bemor_eslatma_ru']:
            setattr(obj, field, request.POST.get(field, ''))
        obj.save()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=obj.patient_id)


@login_required_custom
def doctor_delete_postop(request, doctor_id, pk):
    obj = get_object_or_404(PostOpForm, pk=pk)
    patient_id = obj.patient_id
    obj.delete()
    return redirect('doctor_patient_detail', doctor_id=doctor_id, patient_id=patient_id)


# ─── RECEPTIONIST DOCUMENTS ───

def _doc_entry(form_type, label, obj):
    """Helper to build a document list entry dict."""
    return {
        'type': form_type,
        'label': label,
        'id': obj.pk,
        'protocol': obj.protocol_number,
        'patient_name': f"{obj.patient.name} {obj.patient.surname}",
        'doctor_name': obj.doctor.surname if obj.doctor else '',
        'date': obj.created_at.strftime('%d.%m.%Y %H:%M'),
        'date_raw': obj.created_at,
    }


@login_required_custom
@role_required('receptionist')
def receptionist_reference_indicators(request):
    notif_count, proc_count = _get_receptionist_badge_counts()
    return render(request, 'receptionist/reference_indicators.html', {
        'notif_count': notif_count, 'proc_count': proc_count,
    })


@login_required_custom
@role_required('receptionist', 'admin')
def receptionist_documents(request):
    notif_count, proc_count = _get_receptionist_badge_counts()
    recent_docs = []
    for e in EyeExamForm.objects.select_related('patient', 'doctor').order_by('-created_at')[:10]:
        recent_docs.append(_doc_entry('eye_exam', "Ko'z ko'rigi", e))
    for d in DischargeForm.objects.select_related('patient', 'doctor').order_by('-created_at')[:10]:
        recent_docs.append(_doc_entry('discharge', "Ko'chirma", d))
    for l in LasikForm.objects.select_related('patient', 'doctor').order_by('-created_at')[:10]:
        recent_docs.append(_doc_entry('lasik', 'LASIK', l))
    for s in StrabismusForm.objects.select_related('patient', 'doctor').order_by('-created_at')[:10]:
        recent_docs.append(_doc_entry('strabismus', "Ko'squlilik", s))
    for p in PostOpForm.objects.select_related('patient', 'doctor').order_by('-created_at')[:10]:
        recent_docs.append(_doc_entry('postop', "Operatsionnaya ko'chirma", p))
    # Tibbiy yozuv = all visits where a doctor was assigned (any status except waiting)
    for v in Visit.objects.select_related('patient', 'doctor').filter(
        doctor__isnull=False,
        status__in=['in_progress', 'done', 'has_procedure']
    ).order_by('-created_at')[:30]:
        recent_docs.append({
            'type': 'tibbiy_yozuv', 'label': 'Tibbiy yozuv',
            'id': v.pk, 'protocol': '',
            'patient_name': f"{v.patient.name} {v.patient.surname}",
            'doctor_name': v.doctor.surname if v.doctor else '',
            'date': v.created_at.strftime('%d.%m.%Y %H:%M'),
            'date_raw': v.created_at,
        })
    recent_docs.sort(key=lambda x: x['date_raw'], reverse=True)
    recent_docs = recent_docs[:50]
    return render(request, 'receptionist/documents.html', {
        'notif_count': notif_count,
        'proc_count': proc_count,
        'recent_docs': recent_docs,
    })


@login_required_custom
@role_required('receptionist', 'admin')
def receptionist_documents_search(request):
    query = request.GET.get('q', '').strip()
    results = []
    if query:
        # Search by patient phone or name
        phone = normalize_phone(query)
        patients = Patient.objects.filter(
            Q(phone__icontains=phone) |
            Q(name__icontains=query) |
            Q(surname__icontains=query)
        ).distinct()
        for patient in patients:
            pname = f"{patient.name} {patient.surname}"
            for e in patient.eye_exam_forms.select_related('doctor').order_by('created_at'):
                results.append({'type': 'eye_exam', 'label': "Ko'z ko'rigi",
                               'id': e.pk, 'protocol': e.protocol_number,
                               'patient_name': pname,
                               'doctor_name': e.doctor.surname if e.doctor else '',
                               'date': e.created_at.strftime('%d.%m.%Y %H:%M'),
                               'date_raw': e.created_at.isoformat()})
            for d in patient.discharge_forms.select_related('doctor').order_by('created_at'):
                results.append({'type': 'discharge', 'label': "Ko'chirma",
                               'id': d.pk, 'protocol': d.protocol_number,
                               'patient_name': pname,
                               'doctor_name': d.doctor.surname if d.doctor else '',
                               'date': d.created_at.strftime('%d.%m.%Y %H:%M'),
                               'date_raw': d.created_at.isoformat()})
            for l in patient.lasik_forms.select_related('doctor').order_by('created_at'):
                results.append({'type': 'lasik', 'label': 'LASIK',
                               'id': l.pk, 'protocol': l.protocol_number,
                               'patient_name': pname,
                               'doctor_name': l.doctor.surname if l.doctor else '',
                               'date': l.created_at.strftime('%d.%m.%Y %H:%M'),
                               'date_raw': l.created_at.isoformat()})
            for s in patient.strabismus_forms.select_related('doctor').order_by('created_at'):
                results.append({'type': 'strabismus', 'label': "Ko'squlilik",
                               'id': s.pk, 'protocol': s.protocol_number,
                               'patient_name': pname,
                               'doctor_name': s.doctor.surname if s.doctor else '',
                               'date': s.created_at.strftime('%d.%m.%Y %H:%M'),
                               'date_raw': s.created_at.isoformat()})
            for p in patient.postop_forms.select_related('doctor').order_by('created_at'):
                results.append({'type': 'postop', 'label': "Operatsionnaya ko'chirma",
                               'id': p.pk, 'protocol': p.protocol_number,
                               'patient_name': pname,
                               'doctor_name': p.doctor.surname if p.doctor else '',
                               'date': p.created_at.strftime('%d.%m.%Y %H:%M'),
                               'date_raw': p.created_at.isoformat()})
            for v in patient.visits.select_related('doctor').order_by('created_at'):
                results.append({'type': 'tibbiy_yozuv', 'label': 'Tibbiy yozuv',
                               'id': v.pk, 'protocol': '',
                               'patient_name': pname,
                               'doctor_name': v.doctor.surname if v.doctor else '',
                               'date': v.created_at.strftime('%d.%m.%Y %H:%M'),
                               'date_raw': v.created_at.isoformat()})

        # If no phone match, also try protocol number search
        if not results:
            for Model, form_type, label in [
                (EyeExamForm, 'eye_exam', "Ko'z ko'rigi"),
                (DischargeForm, 'discharge', "Ko'chirma"),
                (LasikForm, 'lasik', 'LASIK'),
                (StrabismusForm, 'strabismus', "Ko'squlilik"),
                (PostOpForm, 'postop', "Operatsionnaya ko'chirma"),
            ]:
                for obj in Model.objects.filter(protocol_number__icontains=query).select_related('patient', 'doctor').order_by('created_at')[:10]:
                    results.append({
                        'type': form_type, 'label': label,
                        'id': obj.pk, 'protocol': obj.protocol_number,
                        'patient_name': f"{obj.patient.name} {obj.patient.surname}",
                        'doctor_name': obj.doctor.surname if obj.doctor else '',
                        'date': obj.created_at.strftime('%d.%m.%Y %H:%M'),
                        'date_raw': obj.created_at.isoformat(),
                    })

        # Sort all results from earliest to latest
        results.sort(key=lambda x: x.get('date_raw', ''))
    return JsonResponse({'results': results})


@login_required_custom
@role_required('receptionist', 'admin')
def receptionist_document_view(request, form_type, pk):
    notif_count, proc_count = _get_receptionist_badge_counts()
    type_map = {
        'eye_exam': (EyeExamForm, 'receptionist/document_eye_exam.html'),
        'discharge': (DischargeForm, 'receptionist/document_discharge.html'),
        'lasik': (LasikForm, 'receptionist/document_lasik.html'),
        'strabismus': (StrabismusForm, 'receptionist/document_strabismus.html'),
        'postop': (PostOpForm, 'receptionist/document_postop.html'),
        'tibbiy_yozuv': (Visit, 'receptionist/document_tibbiy_yozuv.html'),
    }
    if form_type not in type_map:
        from django.http import Http404
        raise Http404
    Model, template = type_map[form_type]
    if form_type == 'tibbiy_yozuv':
        doc = get_object_or_404(
            Model.objects.select_related('patient', 'doctor')
                .prefetch_related('refractions', 'krts', 'iops', 'visual_acuities', 'visit_attachments'),
            pk=pk,
        )
    else:
        doc = get_object_or_404(Model.objects.select_related('patient', 'doctor'), pk=pk)
    return render(request, template, {'doc': doc, 'notif_count': notif_count, 'proc_count': proc_count})


@login_required_custom
@role_required('receptionist', 'admin', 'doctor')
def receptionist_tibbiy_yozuv_print(request, visit_id):
    doc = get_object_or_404(
        Visit.objects.select_related('patient', 'doctor')
            .prefetch_related('refractions', 'krts', 'iops', 'visual_acuities', 'visit_attachments'),
        pk=visit_id,
    )
    show_files = request.GET.get('files', '1') != '0'
    return render(request, 'receptionist/tibbiy_yozuv_print.html', {'doc': doc, 'show_files': show_files})


def visit_share_view(request, token):
    """Public shareable tibbiy yozuv — no login required."""
    doc = get_object_or_404(
        Visit.objects.select_related('patient', 'doctor')
            .prefetch_related('refractions', 'krts', 'iops', 'visual_acuities', 'visit_attachments'),
        share_token=token,
    )
    return render(request, 'receptionist/tibbiy_yozuv_print.html', {'doc': doc, 'show_files': True})


@login_required_custom
@role_required('receptionist', 'admin')
def receptionist_add_attachment(request, visit_id):
    if request.method != 'POST':
        from django.http import HttpResponseNotAllowed
        return HttpResponseNotAllowed(['POST'])
    visit = get_object_or_404(Visit, pk=visit_id)
    title = request.POST.get('title', '').strip()
    files = request.FILES.getlist('files')
    saved = []
    for f in files:
        att = VisitAttachment.objects.create(visit=visit, file=f, title=title)
        saved.append({'url': att.file.url, 'name': f.name, 'title': att.title})
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'files': saved})
    return redirect(request.POST.get('next', reverse('receptionist_panel')))


# ─── LAB: Boshqarish ───

@login_required_custom
@role_required('laboratory')
def lab_boshqarish(request):
    services = LabServiceTemplate.objects.all()
    return render(request, 'lab/boshqarish.html', {'services': services})


@login_required_custom
@role_required('laboratory')
def lab_service_add(request):
    if request.method == 'POST':
        number = request.POST.get('number', '').strip()
        name = request.POST.get('name', '').strip()
        price = request.POST.get('price', '0').strip() or '0'
        ref = request.POST.get('reference_value', '').strip()
        if number and name:
            LabServiceTemplate.objects.get_or_create(
                number=int(number),
                defaults={'name': name, 'price': int(price), 'reference_value': ref},
            )
    return redirect('lab_boshqarish')


@login_required_custom
@role_required('laboratory')
def lab_service_edit(request, pk):
    svc = get_object_or_404(LabServiceTemplate, pk=pk)
    if request.method == 'POST':
        svc.number = int(request.POST.get('number', svc.number))
        svc.name = request.POST.get('name', svc.name).strip()
        svc.price = int(request.POST.get('price', svc.price) or 0)
        svc.reference_value = request.POST.get('reference_value', svc.reference_value).strip()
        svc.save()
    return redirect('lab_boshqarish')


@login_required_custom
@role_required('laboratory')
def lab_service_delete(request, pk):
    svc = get_object_or_404(LabServiceTemplate, pk=pk)
    if request.method == 'POST':
        svc.delete()
    return redirect('lab_boshqarish')


# ─── ANALYTICS: Glasses ───

@role_required('admin')
def analytics_glasses(request):
    period = request.GET.get('period', 'month')
    date_from = _get_date_filter(period)

    sales_qs = OptikaSale.objects.all()
    if date_from:
        sales_qs = sales_qs.filter(created_at__gte=date_from)

    total_revenue = sum(s.total for s in sales_qs)
    total_sales_count = sales_qs.count()

    return render(request, 'admin/analytics_glasses.html', {
        'period': period,
        'total_revenue': total_revenue,
        'total_sales_count': total_sales_count,
        'sales': sales_qs.select_related('patient')[:200],
    })


# ─── OPTIKA PANEL ───

@login_required_custom
@role_required('seller')
def optika_panel(request):
    query = request.GET.get('q', '').strip()
    if query:
        phone = normalize_phone(query)
        patients = Patient.objects.filter(
            Q(name__icontains=query) | Q(surname__icontains=query) | Q(phone__icontains=phone)
        ).distinct().order_by('surname', 'name')
    else:
        patients = Patient.objects.all().order_by('-created_at')[:100]
    return render(request, 'seller/optika_panel.html', {'patients': patients, 'query': query})


@login_required_custom
@role_required('seller')
def optika_patient_detail(request, patient_id):
    patient = get_object_or_404(Patient, pk=patient_id)
    prescription, _ = OptikaPatientPrescription.objects.get_or_create(patient=patient)
    # Pre-fill from latest doctor visit if prescription is empty
    if not any([prescription.retsept_uzoq_od, prescription.retsept_uzoq_os,
                prescription.retsept_yaqin_od, prescription.retsept_yaqin_os,
                prescription.retsept_kontakt_od, prescription.retsept_kontakt_os]):
        latest_visit = Visit.objects.filter(patient=patient, doctor__isnull=False).order_by('-created_at').first()
        if latest_visit:
            prescription.retsept_uzoq_od = latest_visit.retsept_uzoq_od or ''
            prescription.retsept_uzoq_os = latest_visit.retsept_uzoq_os or ''
            prescription.retsept_yaqin_od = latest_visit.retsept_yaqin_od or ''
            prescription.retsept_yaqin_os = latest_visit.retsept_yaqin_os or ''
            prescription.retsept_kontakt_od = latest_visit.retsept_kontakt_od or ''
            prescription.retsept_kontakt_os = latest_visit.retsept_kontakt_os or ''
    return render(request, 'seller/optika_patient_detail.html', {'patient': patient, 'prescription': prescription})


@login_required_custom
@role_required('seller')
def optika_save_prescription(request, patient_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    patient = get_object_or_404(Patient, pk=patient_id)
    prescription, _ = OptikaPatientPrescription.objects.get_or_create(patient=patient)
    prescription.retsept_uzoq_od = request.POST.get('retsept_uzoq_od', '')
    prescription.retsept_uzoq_os = request.POST.get('retsept_uzoq_os', '')
    prescription.retsept_yaqin_od = request.POST.get('retsept_yaqin_od', '')
    prescription.retsept_yaqin_os = request.POST.get('retsept_yaqin_os', '')
    prescription.retsept_kontakt_od = request.POST.get('retsept_kontakt_od', '')
    prescription.retsept_kontakt_os = request.POST.get('retsept_kontakt_os', '')
    prescription.save()
    return JsonResponse({'ok': True})


@login_required_custom
@role_required('seller')
def optika_sell(request, patient_id):
    if request.method != 'POST':
        return redirect('optika_patient_detail', patient_id=patient_id)
    patient = get_object_or_404(Patient, pk=patient_id)
    names = request.POST.getlist('item_name[]')
    prices = request.POST.getlist('item_price[]')
    items = []
    for name, price in zip(names, prices):
        name = name.strip()
        try:
            price = int(str(price).strip().replace(' ', '') or 0)
        except (ValueError, TypeError):
            price = 0
        if name:
            items.append({'name': name, 'price': price})
    total = sum(i['price'] for i in items)
    sale = OptikaSale.objects.create(patient=patient, items=items, total=total)
    return redirect('optika_check', sale_id=sale.pk)


@login_required_custom
@role_required('seller')
def optika_check(request, sale_id):
    sale = get_object_or_404(OptikaSale.objects.select_related('patient'), pk=sale_id)
    return render(request, 'seller/optika_check.html', {'sale': sale})


@login_required_custom
@role_required('seller')
def optika_hisobot(request):
    period = request.GET.get('period', 'month')
    date_from = _get_date_filter(period)
    sales_qs = OptikaSale.objects.select_related('patient').all()
    if date_from:
        sales_qs = sales_qs.filter(created_at__gte=date_from)
    total_revenue = sum(s.total for s in sales_qs)
    total_count = sales_qs.count()
    return render(request, 'seller/optika_hisobot.html', {
        'period': period,
        'total_revenue': total_revenue,
        'total_count': total_count,
        'sales': sales_qs,
    })
